from io import BytesIO

from django.core.exceptions import ValidationError
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook

from apps.common.validators import (
    validate_aadhaar,
    validate_bank_account,
    validate_ifsc,
    validate_mobile,
    validate_pan,
    validate_uan,
)
from apps.company.models import Branch, Department, Designation

from .models import Employee, EmploymentStatus, EmploymentType

EXPORT_HEADERS = [
    'Employee Code',
    'First Name',
    'Last Name',
    'Email',
    'Mobile',
    'Branch Code',
    'Department Code',
    'Designation Code',
    'Date of Joining',
    'Basic Salary',
    'PAN',
    'Aadhaar',
    'UAN',
    'ESIC Number',
    'PF Eligible',
    'ESI Eligible',
    'Employment Type',
    'Employment Status',
    'Bank Name',
    'Account Number',
    'IFSC',
]


def _workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_employees(queryset, filename='employees_export.xlsx'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Employees'
    sheet.append(EXPORT_HEADERS)

    for employee in queryset.select_related('branch', 'department', 'designation'):
        sheet.append([
            employee.employee_code,
            employee.first_name,
            employee.last_name,
            employee.email,
            employee.mobile,
            employee.branch.code if employee.branch else '',
            employee.department.code if employee.department else '',
            employee.designation.code if employee.designation else '',
            employee.date_of_joining.isoformat() if employee.date_of_joining else '',
            float(employee.basic_salary),
            employee.pan,
            employee.aadhaar,
            employee.uan,
            employee.esic_number,
            'Yes' if employee.pf_eligible else 'No',
            'Yes' if employee.esi_eligible else 'No',
            employee.employment_type,
            employee.employment_status,
            employee.bank_name,
            employee.bank_account_number,
            employee.ifsc_code,
        ])

    return _workbook_response(workbook, filename)


def export_employee_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Employees'
    sheet.append(EXPORT_HEADERS)
    sheet.append([
        '',
        'John',
        'Doe',
        'john@example.com',
        '9876543210',
        'HO',
        'FIN',
        'MGR',
        '2026-01-01',
        25000,
        'ABCDE1234F',
        '123456789012',
        '',
        '',
        'Yes',
        'No',
        'permanent',
        'active',
        'State Bank',
        '1234567890',
        'SBIN0001234',
    ])
    return _workbook_response(workbook, 'employee_import_template.xlsx')


def export_import_errors(errors):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Import Errors'
    sheet.append(['Row', 'Error'])
    for error in errors:
        if ': ' in error:
            row_label, message = error.split(': ', 1)
            sheet.append([row_label, message])
        else:
            sheet.append(['', error])
    return _workbook_response(workbook, 'employee_import_errors.xlsx')


def _as_bool(value, default=False):
    if value is None or value == '':
        return default
    return str(value).strip().lower() in ('yes', 'true', '1', 'y')


def _validate_optional_fields(pan, aadhaar, uan, mobile, ifsc, account):
    if pan:
        validate_pan(pan)
    if aadhaar:
        validate_aadhaar(aadhaar)
    if uan:
        validate_uan(uan)
    if mobile:
        validate_mobile(mobile)
    if ifsc:
        validate_ifsc(ifsc)
    if account:
        validate_bank_account(account)


def import_employees(company, uploaded_file, user):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    created = 0
    skipped = 0
    errors = []

    for index, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        try:
            code = str(row[0]).strip().upper() if row[0] else ''
            first_name = str(row[1]).strip() if row[1] else ''
            if not first_name:
                raise ValueError('First name is required.')
            if not row[8]:
                raise ValueError('Date of joining is required.')
            if row[9] in (None, ''):
                raise ValueError('Basic salary is required.')

            if code and Employee.all_objects.filter(company=company, employee_code__iexact=code).exists():
                skipped += 1
                errors.append(f'Row {index}: Skipped duplicate employee code {code}.')
                continue

            branch = None
            department = None
            designation = None
            if row[5]:
                branch = Branch.objects.filter(company=company, code__iexact=str(row[5]).strip()).first()
                if branch is None:
                    raise ValueError(f'Unknown branch code "{row[5]}".')
            if row[6]:
                department = Department.objects.filter(company=company, code__iexact=str(row[6]).strip()).first()
                if department is None:
                    raise ValueError(f'Unknown department code "{row[6]}".')
            if row[7]:
                designation = Designation.objects.filter(company=company, code__iexact=str(row[7]).strip()).first()
                if designation is None:
                    raise ValueError(f'Unknown designation code "{row[7]}".')

            pan = str(row[10]).strip().upper() if row[10] else ''
            aadhaar = ''.join(str(row[11]).split()) if row[11] else ''
            uan = ''.join(str(row[12]).split()) if row[12] else ''
            mobile = ''.join(str(row[4]).split()) if row[4] else ''
            ifsc = str(row[20]).strip().upper() if len(row) > 20 and row[20] else ''
            account = ''.join(str(row[19]).split()) if len(row) > 19 and row[19] else ''

            _validate_optional_fields(pan, aadhaar, uan, mobile, ifsc, account)

            employment_type = str(row[16]).strip().lower() if len(row) > 16 and row[16] else EmploymentType.PERMANENT
            employment_status = str(row[17]).strip().lower() if len(row) > 17 and row[17] else EmploymentStatus.ACTIVE
            if employment_type not in EmploymentType.values:
                raise ValueError(f'Invalid employment type "{employment_type}".')
            if employment_status not in EmploymentStatus.values:
                raise ValueError(f'Invalid employment status "{employment_status}".')

            employee_kwargs = {
                'company': company,
                'branch': branch,
                'department': department,
                'designation': designation,
                'first_name': first_name,
                'last_name': str(row[2]).strip() if row[2] else '',
                'email': str(row[3]).strip().lower() if row[3] else '',
                'mobile': mobile,
                'date_of_joining': row[8],
                'basic_salary': row[9],
                'pan': pan,
                'aadhaar': aadhaar,
                'uan': uan,
                'esic_number': str(row[13]).strip() if row[13] else '',
                'pf_eligible': _as_bool(row[14], default=True),
                'esi_eligible': _as_bool(row[15], default=False),
                'employment_type': employment_type,
                'employment_status': employment_status,
                'bank_name': str(row[18]).strip() if len(row) > 18 and row[18] else '',
                'bank_account_number': account,
                'ifsc_code': ifsc,
                'auto_generate_code': not bool(code),
                'created_by': user,
                'updated_by': user,
            }
            if code:
                employee_kwargs['employee_code'] = code

            Employee.objects.create(**employee_kwargs)
            created += 1
        except (ValidationError, ValueError, TypeError) as exc:
            if hasattr(exc, 'messages'):
                message = '; '.join(exc.messages)
            else:
                message = str(exc)
            errors.append(f'Row {index}: {message}')
        except Exception as exc:
            errors.append(f'Row {index}: {exc}')

    return created, skipped, errors
