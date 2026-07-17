"""Sprint 9.2 ESI compliance tests."""

from datetime import date
from decimal import Decimal
from pathlib import Path

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.exceptions import ValidationError
from django.test import TestCase
from django.urls import reverse

from apps.clients.models import Client
from apps.company.models import Company
from apps.compliance.models import EmployeeESIProfile, ESIRuleSet, PayrollESIResult
from apps.compliance.services.esi_engine import (
    calculate_esi,
    get_esi_contribution_period,
)
from apps.compliance.services.esi_export import iter_esi_contribution_rows
from apps.compliance.services.esi_reports import (
    export_esi_monthly_summary,
    export_esi_register,
    export_missing_ip,
)
from apps.compliance.services.esi_rules import get_esi_rule_for_date, seed_default_esi_rule_set
from apps.compliance.services.pf_rules import seed_default_pf_rule_set
from apps.compliance.services.validator import (
    assert_no_duplicate_esi_ip,
    validate_employee_esi_for_payroll,
    validate_esi_ip_value,
)
from apps.employee.models import Employee
from apps.payroll.models import (
    CalculationType,
    EmployeeSalaryAssignment,
    PayrollResult,
    PayrollRunStatus,
    SalaryComponent,
    SalaryStructure,
    SalaryStructureLine,
)
from apps.payroll.seed import seed_standard_components
from apps.payroll.services.exceptions import LockedRunError
from apps.payroll.services.locking import LockedRunMutationError
from apps.payroll.services.payroll_engine import calculate_run, create_period, create_run
from apps.payroll.services.snapshot import clear_run_results


class ESIComplianceTestMixin:
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='esiadmin',
            password='TestPassword123!',
            is_staff=True,
        )
        for codename in (
            'view_payrollrun',
            'export_esiregister',
            'export_esicontribution',
            'export_pfregister',
            'export_ecr',
        ):
            try:
                perm = Permission.objects.get(codename=codename)
                self.user.user_permissions.add(perm)
            except Permission.DoesNotExist:
                pass

        self.client_record = Client.objects.create(
            client_code='ESICLI',
            client_name='ESI Client',
            mobile='9876543210',
            address_line_1='Naidupet',
            city='Naidupet',
            state='Andhra Pradesh',
            pincode='524126',
        )
        self.company = Company.objects.create(
            client=self.client_record,
            company_name='ESI Co',
        )
        self.employee = Employee.objects.create(
            company=self.company,
            employee_code='ESI1001',
            first_name='Ravi',
            last_name='Kumar',
            date_of_joining=date(2024, 1, 1),
            basic_salary=Decimal('15000.00'),
            auto_generate_code=False,
            pf_eligible=True,
            esi_eligible=True,
            uan='123456789012',
        )
        self.pf_rule = seed_default_pf_rule_set()
        self.rule = seed_default_esi_rule_set()
        EmployeeESIProfile.objects.create(
            employee=self.employee,
            ip_number='1234567890',
            is_esi_applicable=True,
            joining_esi_date=date(2024, 1, 1),
        )

    def _period(self, start=None, end=None):
        return type('P', (), {
            'start_date': start or date(2026, 7, 1),
            'end_date': end or date(2026, 7, 31),
        })()

    def _rows(self, amount, *, esi=True):
        return [{
            'component_code': 'BASIC',
            'amount': Decimal(amount),
            'esi_applicable': esi,
            'pf_applicable': True,
        }]


class ESIRuleResolutionTests(ESIComplianceTestMixin, TestCase):
    def test_seed_default_rates(self):
        self.assertEqual(self.rule.eligibility_wage_limit, Decimal('21000.00'))
        self.assertEqual(self.rule.employee_rate, Decimal('0.0075'))
        self.assertEqual(self.rule.employer_rate, Decimal('0.0325'))
        self.assertEqual(self.rule.daily_wage_exemption_limit, Decimal('176.00'))

    def test_effective_date_resolution(self):
        resolved = get_esi_rule_for_date(date(2026, 7, 31))
        self.assertEqual(resolved.code, self.rule.code)

    def test_no_rule_before_effective(self):
        with self.assertRaises(ValidationError):
            get_esi_rule_for_date(date(2020, 1, 1))

    def test_overlapping_active_rules_rejected(self):
        with self.assertRaises(ValidationError):
            ESIRuleSet.objects.create(
                code='IN-ESI-OVERLAP',
                name='Overlap',
                effective_from=date(2024, 6, 1),
                effective_to=None,
                is_active=True,
            )

    def test_rule_change_by_date(self):
        self.rule.effective_to = date(2025, 3, 31)
        self.rule.save()
        new_rule = ESIRuleSet.objects.create(
            code='IN-ESI-2025',
            name='India ESI 2025',
            effective_from=date(2025, 4, 1),
            effective_to=None,
            is_active=True,
            eligibility_wage_limit=Decimal('25000.00'),
            employee_rate=Decimal('0.0075'),
            employer_rate=Decimal('0.0325'),
            daily_wage_exemption_limit=Decimal('176.00'),
        )
        self.assertEqual(get_esi_rule_for_date(date(2025, 2, 1)).code, self.rule.code)
        self.assertEqual(get_esi_rule_for_date(date(2025, 5, 1)).code, new_rule.code)
        self.assertEqual(get_esi_rule_for_date(date(2025, 5, 1)).eligibility_wage_limit, Decimal('25000.00'))

    def test_contribution_period_bounds(self):
        self.assertEqual(
            get_esi_contribution_period(date(2026, 7, 15)),
            (date(2026, 4, 1), date(2026, 9, 30)),
        )
        self.assertEqual(
            get_esi_contribution_period(date(2026, 11, 1)),
            (date(2026, 10, 1), date(2027, 3, 31)),
        )
        self.assertEqual(
            get_esi_contribution_period(date(2027, 2, 1)),
            (date(2026, 10, 1), date(2027, 3, 31)),
        )


class ESIEngineCalculationTests(ESIComplianceTestMixin, TestCase):
    def test_below_ceiling(self):
        esi = calculate_esi(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('20000.00'),
            rule_set=self.rule,
        )
        self.assertTrue(esi.eligible)
        self.assertEqual(esi.esi_wages, Decimal('20000.00'))
        self.assertEqual(esi.employee_esi, Decimal('150.00'))  # 20000 * 0.75%
        self.assertEqual(esi.employer_esi, Decimal('650.00'))  # 20000 * 3.25%
        self.assertFalse(esi.above_wage_limit)

    def test_above_ceiling_at_first_eligibility(self):
        esi = calculate_esi(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('25000.00'),
            rule_set=self.rule,
            update_continuity=False,
        )
        self.assertFalse(esi.eligible)
        self.assertTrue(esi.above_wage_limit)
        self.assertEqual(esi.employee_esi, Decimal('0.00'))
        self.assertEqual(esi.employer_esi, Decimal('0.00'))
        self.assertEqual(esi.eligibility_notes, 'above_wage_limit_no_continuity')

    def test_crossing_ceiling_during_contribution_period(self):
        # First month below ceiling → covered
        first = calculate_esi(
            employee=self.employee,
            period=self._period(date(2026, 4, 1), date(2026, 4, 30)),
            earning_rows=self._rows('20000.00'),
            rule_set=self.rule,
        )
        self.assertTrue(first.eligible)
        profile = self.employee.esi_profile
        profile.refresh_from_db()
        self.assertEqual(profile.covered_period_start, date(2026, 4, 1))

        # Later month above ceiling → still eligible via continuity
        second = calculate_esi(
            employee=self.employee,
            period=self._period(date(2026, 7, 1), date(2026, 7, 31)),
            earning_rows=self._rows('25000.00'),
            rule_set=self.rule,
        )
        self.assertTrue(second.eligible)
        self.assertTrue(second.above_wage_limit)
        self.assertTrue(second.continuity_applied)
        self.assertEqual(second.employee_esi, Decimal('187.50'))  # 25000 * 0.75%
        self.assertEqual(second.employer_esi, Decimal('812.50'))  # 25000 * 3.25%

    def test_daily_wage_exemption(self):
        # Avg daily = 5000/31 ≈ 161.29 ≤ 176 → EE exempt
        esi = calculate_esi(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('5000.00'),
            rule_set=self.rule,
            calendar_days=Decimal('31'),
        )
        self.assertTrue(esi.eligible)
        self.assertTrue(esi.daily_wage_exemption)
        self.assertEqual(esi.employee_esi, Decimal('0.00'))
        self.assertEqual(esi.employer_esi, Decimal('162.50'))  # 5000 * 3.25%

    def test_mid_month_join(self):
        self.employee.date_of_joining = date(2026, 7, 16)
        self.employee.save(update_fields=['date_of_joining'])
        profile = self.employee.esi_profile
        profile.joining_esi_date = date(2026, 7, 16)
        profile.save(update_fields=['joining_esi_date', 'updated_at'])
        # Prorated wages already reflected in earning_rows
        esi = calculate_esi(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('10000.00'),
            rule_set=self.rule,
        )
        self.assertTrue(esi.eligible)
        self.assertEqual(esi.employee_esi, Decimal('75.00'))

    def test_exit_before_period(self):
        profile = self.employee.esi_profile
        profile.exit_esi_date = date(2026, 6, 30)
        profile.save(update_fields=['exit_esi_date', 'updated_at'])
        esi = calculate_esi(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('15000.00'),
            rule_set=self.rule,
        )
        self.assertFalse(esi.eligible)
        self.assertEqual(esi.eligibility_notes, 'exited_before_period')

    def test_zero_esi_wages(self):
        esi = calculate_esi(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('0.00'),
            rule_set=self.rule,
        )
        self.assertTrue(esi.eligible)
        self.assertEqual(esi.employee_esi, Decimal('0.00'))
        self.assertEqual(esi.employer_esi, Decimal('0.00'))
        self.assertIn('zero_esi_wages', esi.eligibility_notes)

    def test_non_esi_employee(self):
        profile = self.employee.esi_profile
        profile.is_esi_applicable = False
        profile.save(update_fields=['is_esi_applicable', 'updated_at'])
        esi = calculate_esi(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('15000.00'),
            rule_set=self.rule,
        )
        self.assertFalse(esi.eligible)
        self.assertEqual(esi.employee_esi, Decimal('0.00'))

    def test_missing_ip_flagged(self):
        profile = self.employee.esi_profile
        profile.ip_number = ''
        profile.save(update_fields=['ip_number', 'updated_at'])
        esi = calculate_esi(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('15000.00'),
            rule_set=self.rule,
        )
        self.assertTrue(esi.eligible)
        self.assertTrue(esi.missing_ip_number)
        self.assertIn('missing_ip_number', esi.eligibility_notes)


class ESIValidationTests(ESIComplianceTestMixin, TestCase):
    def test_ip_format(self):
        self.assertEqual(validate_esi_ip_value('1234567890'), '1234567890')
        with self.assertRaises(ValidationError):
            validate_esi_ip_value('ABC')

    def test_missing_ip_when_required(self):
        profile = self.employee.esi_profile
        profile.ip_number = ''
        profile.save(update_fields=['ip_number', 'updated_at'])
        errs = validate_employee_esi_for_payroll(self.employee, require_ip=True)
        self.assertTrue(any('Missing ESI IP' in e for e in errs))

    def test_duplicate_ip(self):
        other = Employee.objects.create(
            company=self.company,
            employee_code='ESI1002',
            first_name='Other',
            last_name='Emp',
            date_of_joining=date(2024, 1, 1),
            basic_salary=Decimal('10000.00'),
            auto_generate_code=False,
            esi_eligible=True,
        )
        with self.assertRaises(ValidationError):
            assert_no_duplicate_esi_ip('1234567890')
        with self.assertRaises(ValidationError):
            EmployeeESIProfile.objects.create(
                employee=other,
                ip_number='1234567890',
                is_esi_applicable=True,
            )


class PayrollESIIntegrationTests(ESIComplianceTestMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.period = create_period(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            user=self.user,
        )
        self.run = create_run(period=self.period, user=self.user)
        seed_standard_components(self.company)
        structure = SalaryStructure.objects.create(
            company=self.company,
            code='ESI92',
            name='ESI Structure',
        )
        basic = SalaryComponent.objects.get(company=self.company, component_code='BASIC')
        SalaryStructureLine.objects.create(
            structure=structure,
            component=basic,
            calculation_type=CalculationType.PERCENTAGE,
            percent=Decimal('40.0000'),
            display_order=10,
        )
        EmployeeSalaryAssignment.objects.create(
            employee=self.employee,
            salary_structure=structure,
            effective_from=date(2024, 1, 1),
            gross_salary=Decimal('45000.00'),
            ctc=Decimal('540000.00'),
            created_by=self.user,
            updated_by=self.user,
        )

    def test_calculate_run_persists_esi_and_rule_snapshot(self):
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        self.assertEqual(self.run.esi_rule_set_id, self.rule.pk)
        result = PayrollResult.objects.get(run=self.run, employee=self.employee)
        self.assertTrue(hasattr(result, 'esi_result'))
        esi = result.esi_result
        self.assertTrue(esi.is_eligible)
        self.assertEqual(esi.rule_version, self.rule.code)
        # BASIC 40% of 45000 = 18000
        self.assertEqual(esi.esi_wages, Decimal('18000.00'))
        self.assertEqual(esi.employee_esi, Decimal('135.00'))
        self.assertEqual(esi.employer_esi, Decimal('585.00'))

        # Historical: changing live rule rates does not mutate snapshot
        self.rule.employee_rate = Decimal('0.0100')
        self.rule.save(update_fields=['employee_rate', 'updated_at'])
        esi.refresh_from_db()
        self.assertEqual(esi.employee_esi, Decimal('135.00'))

    def test_unlocked_recalculation(self):
        calculate_run(self.run, user=self.user)
        first_id = PayrollESIResult.objects.get(payroll_result__run=self.run).pk
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        second = PayrollESIResult.objects.get(payroll_result__run=self.run)
        self.assertNotEqual(first_id, second.pk)
        self.assertEqual(second.employee_esi, Decimal('135.00'))

    def test_locked_rejection(self):
        calculate_run(self.run, user=self.user)
        self.run.status = PayrollRunStatus.LOCKED
        self.run.save(update_fields=['status', 'updated_at'])
        with self.assertRaises(LockedRunError):
            calculate_run(self.run, user=self.user)
        with self.assertRaises(LockedRunMutationError):
            clear_run_results(self.run)

    def test_hub_and_exports(self):
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.client.force_login(self.user)
        hub = self.client.get(reverse('compliance:hub'))
        self.assertEqual(hub.status_code, 200)
        self.assertContains(hub, 'ESI Register')

        reg = self.client.get(
            reverse('compliance:esi_report_export', kwargs={
                'run_id': self.run.pk,
                'report_key': 'esi_register',
            })
        )
        self.assertEqual(reg.status_code, 200)
        self.assertIn(
            'spreadsheetml',
            reg['Content-Type'],
        )

        missing = export_missing_ip(self.run)
        self.assertEqual(missing.status_code, 200)

        summary = export_esi_monthly_summary(self.run)
        self.assertEqual(summary.status_code, 200)

        rows = list(iter_esi_contribution_rows(self.run, require_ip=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['employee_esi'], Decimal('135.00'))
        # Report reconcile with snapshot
        snap = PayrollESIResult.objects.get(payroll_result__run=self.run)
        self.assertEqual(rows[0]['esi_wages'], snap.esi_wages)
        self.assertEqual(rows[0]['employer_esi'], snap.employer_esi)

        contrib = self.client.get(
            reverse('compliance:esi_contribution_export', kwargs={'run_id': self.run.pk})
        )
        self.assertEqual(contrib.status_code, 200)

    def test_missing_ip_report_not_silent(self):
        from io import BytesIO

        from openpyxl import load_workbook

        profile = self.employee.esi_profile
        profile.ip_number = ''
        profile.save(update_fields=['ip_number', 'updated_at'])
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        snap = PayrollESIResult.objects.get(payroll_result__run=self.run)
        self.assertTrue(snap.missing_ip_number)
        response = export_missing_ip(self.run)
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        rows = list(wb.active.iter_rows(values_only=True))
        codes = {r[0] for r in rows[1:] if r and r[0]}
        self.assertIn('ESI1001', codes)

    def test_no_eval_in_esi_engine(self):
        engine_path = (
            Path(__file__).resolve().parents[1] / 'services' / 'esi_engine.py'
        )
        source = engine_path.read_text(encoding='utf-8')
        self.assertNotIn('eval(', source)
        self.assertNotIn('exec(', source)
