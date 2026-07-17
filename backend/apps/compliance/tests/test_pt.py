"""Sprint 9.3 Professional Tax compliance tests."""

from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.exceptions import ValidationError
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.clients.models import Client
from apps.company.models import Company
from apps.compliance.models import (
    EmployeePTProfile,
    PayrollPTResult,
    ProfessionalTaxRuleSet,
    ProfessionalTaxSlab,
    PTExemptionType,
)
from apps.compliance.services.esi_rules import seed_default_esi_rule_set
from apps.compliance.services.pf_rules import seed_default_pf_rule_set
from apps.compliance.services.pt_engine import calculate_pt
from apps.compliance.services.pt_export import iter_pt_challan_rows
from apps.compliance.services.pt_reports import (
    export_missing_pt_work_state,
    export_pt_monthly_summary,
    export_pt_register,
)
from apps.compliance.services.pt_rules import (
    AP_SPECIAL_MONTH,
    get_pt_rule_for_state_and_date,
    seed_ap_pt_rule_set,
)
from apps.employee.models import Employee
from apps.payroll.models import (
    CalculationType,
    EmployeeSalaryAssignment,
    PayrollResult,
    PayrollRunStatus,
    SalaryComponent,
    SalaryStructure,
    SalaryStructureLine,
)
from apps.payroll.seed import seed_standard_components
from apps.payroll.services.exceptions import LockedRunError
from apps.payroll.services.locking import LockedRunMutationError
from apps.payroll.services.payroll_engine import calculate_run, create_period, create_run
from apps.payroll.services.snapshot import clear_run_results


class PTComplianceTestMixin:
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='ptadmin',
            password='TestPassword123!',
            is_staff=True,
        )
        for codename in (
            'view_payrollrun',
            'export_ptregister',
            'export_ptchallan',
            'export_esiregister',
            'export_esicontribution',
            'export_pfregister',
            'export_ecr',
        ):
            try:
                perm = Permission.objects.get(codename=codename)
                self.user.user_permissions.add(perm)
            except Permission.DoesNotExist:
                pass

        self.client_record = Client.objects.create(
            client_code='PTCLI',
            client_name='PT Client',
            mobile='9876543210',
            address_line_1='Naidupet',
            city='Naidupet',
            state='Andhra Pradesh',
            pincode='524126',
        )
        self.company = Company.objects.create(
            client=self.client_record,
            company_name='PT Co',
            state='Andhra Pradesh',
        )
        self.employee = Employee.objects.create(
            company=self.company,
            employee_code='PT1001',
            first_name='Ravi',
            last_name='Kumar',
            date_of_joining=date(2024, 1, 1),
            basic_salary=Decimal('15000.00'),
            auto_generate_code=False,
            pf_eligible=True,
            esi_eligible=True,
            uan='123456789012',
        )
        self.pf_rule = seed_default_pf_rule_set()
        self.esi_rule = seed_default_esi_rule_set()
        self.rule = seed_ap_pt_rule_set(force_slabs=True)
        EmployeePTProfile.objects.create(
            employee=self.employee,
            state_code='AP',
            is_applicable=True,
            effective_from=date(2024, 1, 1),
        )

    def _period(self, start=None, end=None):
        return type('P', (), {
            'start_date': start or date(2026, 7, 1),
            'end_date': end or date(2026, 7, 31),
        })()

    def _rows(self, amount, *, pt=True):
        return [{
            'component_code': 'BASIC',
            'amount': Decimal(amount),
            'pt_applicable': pt,
            'pf_applicable': True,
            'esi_applicable': True,
        }]


class PTRuleResolutionTests(PTComplianceTestMixin, TestCase):
    def test_seed_ap_slabs(self):
        slabs = list(self.rule.slabs.order_by('sequence'))
        self.assertEqual(len(slabs), 3)
        self.assertEqual(self.rule.special_month, AP_SPECIAL_MONTH)
        self.assertEqual(self.rule.special_month, 2)
        self.assertEqual(slabs[0].salary_to, Decimal('15000.00'))
        self.assertEqual(slabs[0].tax_amount, Decimal('0.00'))
        self.assertEqual(slabs[1].salary_from, Decimal('15001.00'))
        self.assertEqual(slabs[1].salary_to, Decimal('20000.00'))
        self.assertEqual(slabs[1].tax_amount, Decimal('150.00'))
        self.assertEqual(slabs[2].salary_from, Decimal('20001.00'))
        self.assertIsNone(slabs[2].salary_to)
        self.assertEqual(slabs[2].tax_amount, Decimal('200.00'))
        self.assertEqual(slabs[2].special_month_tax_amount, Decimal('300.00'))

    def test_effective_date_resolution(self):
        resolved = get_pt_rule_for_state_and_date('AP', date(2026, 7, 31))
        self.assertEqual(resolved.pk, self.rule.pk)

    def test_no_rule_before_effective(self):
        with self.assertRaises(ValidationError):
            get_pt_rule_for_state_and_date('AP', date(2020, 1, 1))

    def test_overlapping_active_rules_rejected(self):
        with self.assertRaises(ValidationError):
            ProfessionalTaxRuleSet.objects.create(
                state_code='AP',
                name='AP Overlap',
                effective_from=date(2024, 6, 1),
                effective_to=None,
                is_active=True,
                special_month=2,
            )

    def test_rule_change_by_date(self):
        self.rule.effective_to = date(2025, 3, 31)
        self.rule.save()
        new_rule = ProfessionalTaxRuleSet.objects.create(
            state_code='AP',
            name='AP PT 2025',
            effective_from=date(2025, 4, 1),
            effective_to=None,
            is_active=True,
            special_month=2,
        )
        ProfessionalTaxSlab.objects.create(
            rule_set=new_rule,
            salary_from=Decimal('0.00'),
            salary_to=None,
            tax_amount=Decimal('250.00'),
            special_month_tax_amount=Decimal('350.00'),
            sequence=10,
        )
        self.assertEqual(get_pt_rule_for_state_and_date('AP', date(2025, 2, 1)).pk, self.rule.pk)
        resolved = get_pt_rule_for_state_and_date('AP', date(2025, 5, 1))
        self.assertEqual(resolved.pk, new_rule.pk)
        pt = calculate_pt(
            employee=self.employee,
            period=self._period(date(2025, 5, 1), date(2025, 5, 31)),
            earning_rows=self._rows('25000.00'),
        )
        self.assertEqual(pt.tax_amount, Decimal('250.00'))
        self.assertEqual(pt.rule_set_id, new_rule.pk)


class PTEngineSlabBoundaryTests(PTComplianceTestMixin, TestCase):
    def _calc(self, wages, *, start=None, end=None):
        return calculate_pt(
            employee=self.employee,
            period=self._period(start, end),
            earning_rows=self._rows(wages),
            rule_set=self.rule,
        )

    def test_below_taxable(self):
        pt = self._calc('14999.00')
        self.assertTrue(pt.applicable)
        self.assertEqual(pt.tax_amount, Decimal('0.00'))

    def test_lower_boundary_nil_slab(self):
        pt = self._calc('15000.00')
        self.assertEqual(pt.tax_amount, Decimal('0.00'))

    def test_upper_boundary_nil_to_mid(self):
        pt = self._calc('15001.00')
        self.assertEqual(pt.tax_amount, Decimal('150.00'))

    def test_mid_slab_upper(self):
        pt = self._calc('20000.00')
        self.assertEqual(pt.tax_amount, Decimal('150.00'))

    def test_top_slab_lower(self):
        pt = self._calc('20001.00')
        self.assertEqual(pt.tax_amount, Decimal('200.00'))

    def test_top_slab_mid(self):
        pt = self._calc('45000.00')
        self.assertEqual(pt.tax_amount, Decimal('200.00'))
        self.assertFalse(pt.special_month_applied)

    def test_special_month_february(self):
        pt = self._calc(
            '45000.00',
            start=date(2026, 2, 1),
            end=date(2026, 2, 28),
        )
        self.assertTrue(pt.special_month_applied)
        self.assertEqual(pt.tax_amount, Decimal('300.00'))

    def test_special_month_mid_slab_unchanged(self):
        pt = self._calc(
            '18000.00',
            start=date(2026, 2, 1),
            end=date(2026, 2, 28),
        )
        self.assertTrue(pt.special_month_applied)
        self.assertEqual(pt.tax_amount, Decimal('150.00'))

    def test_no_hardcoded_slabs_in_engine(self):
        engine_path = Path(__file__).resolve().parents[1] / 'services' / 'pt_engine.py'
        source = engine_path.read_text(encoding='utf-8')
        self.assertNotIn('15001', source)
        self.assertNotIn('20001', source)
        self.assertNotIn('Decimal(\'200.00\')', source)
        self.assertNotIn('Decimal(\'300.00\')', source)
        self.assertNotIn('eval(', source)
        self.assertNotIn('exec(', source)


class PTApplicabilityTests(PTComplianceTestMixin, TestCase):
    def test_exempt_employee_keeps_reason(self):
        profile = self.employee.pt_profile
        profile.is_applicable = False
        profile.exemption_type = PTExemptionType.SENIOR_CITIZEN
        profile.exemption_reason = 'Senior citizen exemption'
        profile.save()
        pt = calculate_pt(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('45000.00'),
            rule_set=self.rule,
        )
        self.assertFalse(pt.applicable)
        self.assertEqual(pt.tax_amount, Decimal('0.00'))
        self.assertEqual(pt.exemption_reason, 'Senior citizen exemption')

    def test_missing_profile(self):
        EmployeePTProfile.objects.filter(employee=self.employee).delete()
        # Drop any cached reverse OneToOne on the in-memory employee instance.
        self.employee = Employee.objects.get(pk=self.employee.pk)
        pt = calculate_pt(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('45000.00'),
        )
        self.assertFalse(pt.applicable)
        self.assertTrue(pt.missing_work_state)
        self.assertEqual(pt.exemption_reason, 'missing_pt_profile')
        self.assertEqual(pt.tax_amount, Decimal('0.00'))

    def test_missing_work_state_not_company_fallback(self):
        profile = self.employee.pt_profile
        profile.state_code = ''
        profile.is_applicable = False
        profile.save(update_fields=['state_code', 'is_applicable', 'updated_at'])
        # Force applicable with blank state via update_fields bypass for scenario
        EmployeePTProfile.objects.filter(pk=profile.pk).update(
            is_applicable=True,
            state_code='',
        )
        self.employee.refresh_from_db()
        pt = calculate_pt(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('45000.00'),
            allow_company_fallback=False,
        )
        self.assertFalse(pt.applicable)
        self.assertTrue(pt.missing_work_state)
        self.assertEqual(pt.tax_amount, Decimal('0.00'))

    def test_mid_month_join_uses_prorated_wages(self):
        self.employee.date_of_joining = date(2026, 7, 16)
        self.employee.save(update_fields=['date_of_joining'])
        # Half-month wages at mid slab → still ₹150
        pt = calculate_pt(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('9000.00'),
            rule_set=self.rule,
        )
        self.assertTrue(pt.applicable)
        self.assertEqual(pt.tax_amount, Decimal('0.00'))  # below 15000

        pt2 = calculate_pt(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('16000.00'),
            rule_set=self.rule,
        )
        self.assertEqual(pt2.tax_amount, Decimal('150.00'))

    def test_exit_before_period(self):
        self.employee.date_of_exit = date(2026, 6, 30)
        self.employee.save(update_fields=['date_of_exit'])
        pt = calculate_pt(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('45000.00'),
            rule_set=self.rule,
        )
        self.assertFalse(pt.applicable)
        self.assertEqual(pt.exemption_reason, 'exited_before_period')

    def test_state_transfer(self):
        # Create TS rule for transfer target
        ts = ProfessionalTaxRuleSet.objects.create(
            state_code='TS',
            name='Telangana PT test',
            effective_from=date(2024, 4, 1),
            is_active=True,
            special_month=2,
        )
        ProfessionalTaxSlab.objects.create(
            rule_set=ts,
            salary_from=Decimal('0.00'),
            salary_to=None,
            tax_amount=Decimal('200.00'),
            special_month_tax_amount=Decimal('200.00'),
            sequence=10,
        )
        profile = self.employee.pt_profile
        profile.state_code = 'TS'
        profile.save(update_fields=['state_code', 'updated_at'])
        pt = calculate_pt(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('45000.00'),
        )
        self.assertEqual(pt.state_code, 'TS')
        self.assertEqual(pt.tax_amount, Decimal('200.00'))
        self.assertEqual(pt.rule_set_id, ts.pk)


class PayrollPTIntegrationTests(PTComplianceTestMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.period = create_period(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            user=self.user,
        )
        self.run = create_run(period=self.period, user=self.user)
        seed_standard_components(self.company)
        structure = SalaryStructure.objects.create(
            company=self.company,
            code='PT93',
            name='PT Structure',
        )
        basic = SalaryComponent.objects.get(company=self.company, component_code='BASIC')
        SalaryStructureLine.objects.create(
            structure=structure,
            component=basic,
            calculation_type=CalculationType.PERCENTAGE,
            percent=Decimal('40.0000'),
            display_order=10,
        )
        # Gross 50000 → BASIC 20000 → mid slab ₹150
        EmployeeSalaryAssignment.objects.create(
            employee=self.employee,
            salary_structure=structure,
            effective_from=date(2024, 1, 1),
            gross_salary=Decimal('50000.00'),
            ctc=Decimal('600000.00'),
            created_by=self.user,
            updated_by=self.user,
        )

    def test_calculate_run_persists_pt_and_snapshot(self):
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        self.assertEqual(self.run.pt_rule_set_id, self.rule.pk)
        result = PayrollResult.objects.get(run=self.run, employee=self.employee)
        self.assertTrue(hasattr(result, 'pt_result'))
        pt = result.pt_result
        self.assertEqual(pt.state_code, 'AP')
        self.assertEqual(pt.pt_wages, Decimal('20000.00'))
        self.assertEqual(pt.tax_amount, Decimal('150.00'))
        # Historical: changing live slab does not mutate snapshot
        top = self.rule.slabs.get(sequence=20)
        top.tax_amount = Decimal('999.00')
        top.save(update_fields=['tax_amount', 'updated_at'])
        pt.refresh_from_db()
        self.assertEqual(pt.tax_amount, Decimal('150.00'))

    def test_unlocked_recalculation(self):
        calculate_run(self.run, user=self.user)
        first_id = PayrollPTResult.objects.get(payroll_result__run=self.run).pk
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        second = PayrollPTResult.objects.get(payroll_result__run=self.run)
        self.assertNotEqual(first_id, second.pk)
        self.assertEqual(second.tax_amount, Decimal('150.00'))

    def test_locked_rejection(self):
        calculate_run(self.run, user=self.user)
        self.run.status = PayrollRunStatus.LOCKED
        self.run.save(update_fields=['status', 'updated_at'])
        with self.assertRaises(LockedRunError):
            calculate_run(self.run, user=self.user)
        with self.assertRaises(LockedRunMutationError):
            clear_run_results(self.run)

    def test_hub_and_exports_reconcile(self):
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.client.force_login(self.user)
        hub = self.client.get(reverse('compliance:hub'))
        self.assertEqual(hub.status_code, 200)
        self.assertContains(hub, 'PT Register')

        reg = self.client.get(
            reverse('compliance:pt_report_export', kwargs={
                'run_id': self.run.pk,
                'report_key': 'pt_register',
            })
        )
        self.assertEqual(reg.status_code, 200)
        self.assertIn('spreadsheetml', reg['Content-Type'])

        summary = export_pt_monthly_summary(self.run)
        self.assertEqual(summary.status_code, 200)

        rows = list(iter_pt_challan_rows(self.run))
        self.assertEqual(len(rows), 1)
        snap = PayrollPTResult.objects.get(payroll_result__run=self.run)
        self.assertEqual(rows[0]['tax_amount'], snap.tax_amount)
        self.assertEqual(rows[0]['pt_wages'], snap.pt_wages)

        challan = self.client.get(
            reverse('compliance:pt_challan_export', kwargs={'run_id': self.run.pk})
        )
        self.assertEqual(challan.status_code, 200)

        register = export_pt_register(self.run)
        wb = load_workbook(BytesIO(register.content))
        data_rows = list(wb.active.iter_rows(values_only=True))
        self.assertIn('PT1001', {r[0] for r in data_rows[1:] if r and r[0]})

    def test_missing_work_state_report(self):
        EmployeePTProfile.objects.filter(employee=self.employee).delete()
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        snap = PayrollPTResult.objects.get(payroll_result__run=self.run)
        self.assertTrue(snap.calculation_snapshot.get('missing_work_state'))
        self.assertEqual(snap.tax_amount, Decimal('0.00'))
        response = export_missing_pt_work_state(self.run)
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        codes = {r[0] for r in list(wb.active.iter_rows(values_only=True))[1:] if r and r[0]}
        self.assertIn('PT1001', codes)

    def test_february_special_month_in_run(self):
        feb_period = create_period(
            company=self.company,
            month=2,
            year=2026,
            start_date=date(2026, 2, 1),
            end_date=date(2026, 2, 28),
            user=self.user,
        )
        feb_run = create_run(period=feb_period, user=self.user)
        # Raise gross so BASIC (40%) exceeds 20000 → top slab special month ₹300
        assignment = EmployeeSalaryAssignment.objects.get(employee=self.employee)
        assignment.gross_salary = Decimal('60000.00')
        assignment.save(update_fields=['gross_salary', 'updated_at'])
        calculate_run(feb_run, user=self.user)
        pt = PayrollPTResult.objects.get(payroll_result__run=feb_run)
        self.assertEqual(pt.pt_wages, Decimal('24000.00'))
        self.assertEqual(pt.tax_amount, Decimal('300.00'))
        self.assertTrue(pt.calculation_snapshot.get('special_month_applied'))
