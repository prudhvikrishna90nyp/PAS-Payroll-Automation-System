"""Sprint 9.4 TDS / income-tax compliance tests."""

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.exceptions import ValidationError
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.clients.models import Client
from apps.company.models import Company
from apps.compliance.models import (
    EmployeeTaxProfile,
    FinancialYearTaxRule,
    PayrollTDSResult,
    PreviousEmployerIncome,
    TaxDeclaration,
    TaxDeclarationStatus,
    TaxRegime,
    TaxSlab,
)
from apps.compliance.services.esi_rules import seed_default_esi_rule_set
from apps.compliance.services.pf_rules import seed_default_pf_rule_set
from apps.compliance.services.pt_rules import seed_ap_pt_rule_set
from apps.compliance.services.tds_engine import (
    apply_rebate,
    apply_surcharge,
    calculate_tds_for_employee,
    compute_slab_tax,
    resolve_regime,
)
from apps.compliance.services.tds_reports import (
    export_tds_monthly_summary,
    export_tds_register,
)
from apps.compliance.services.tds_rules import (
    financial_year_for_date,
    get_tax_rule_for_fy_regime_and_date,
    regime_change_deadline,
    seed_tds_rule_sets,
)
from apps.employee.models import Employee
from apps.payroll.models import (
    CalculationType,
    EmployeeSalaryAssignment,
    PayrollResult,
    PayrollRunStatus,
    SalaryComponent,
    SalaryStructure,
    SalaryStructureLine,
)
from apps.payroll.seed import seed_standard_components
from apps.payroll.services.locking import LockedRunMutationError
from apps.payroll.services.payroll_engine import calculate_run, create_period, create_run
from apps.payroll.services.snapshot import clear_run_results


class TDSComplianceTestMixin:
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='tdsadmin',
            password='TestPassword123!',
            is_staff=True,
        )
        for codename in (
            'view_payrollrun',
            'export_tdsregister',
            'export_form16',
            'export_ptregister',
            'export_pfregister',
            'export_esiregister',
            'export_ecr',
        ):
            try:
                perm = Permission.objects.get(codename=codename)
                self.user.user_permissions.add(perm)
            except Permission.DoesNotExist:
                pass

        self.client_record = Client.objects.create(
            client_code='TDSCLI',
            client_name='TDS Client',
            mobile='9876543210',
            address_line_1='Hyderabad',
            city='Hyderabad',
            state='Telangana',
            pincode='500001',
        )
        self.company = Company.objects.create(
            client=self.client_record,
            company_name='TDS Co',
            state='Telangana',
        )
        self.employee = Employee.objects.create(
            company=self.company,
            employee_code='TDS1001',
            first_name='Anita',
            last_name='Sharma',
            date_of_joining=date(2024, 1, 1),
            basic_salary=Decimal('50000.00'),
            auto_generate_code=False,
            pf_eligible=True,
            esi_eligible=False,
            uan='123456789012',
        )
        seed_default_pf_rule_set()
        seed_default_esi_rule_set()
        seed_ap_pt_rule_set()
        self.rules = seed_tds_rule_sets(force_slabs=True)
        self.new_rule = FinancialYearTaxRule.objects.get(code='IN-TDS-NEW-2025-26')
        self.old_rule = FinancialYearTaxRule.objects.get(code='IN-TDS-OLD-2025-26')
        EmployeeTaxProfile.objects.create(
            employee=self.employee,
            default_tax_regime=TaxRegime.NEW,
            pan_number='ABCDE1234F',
            effective_from=date(2025, 4, 1),
            is_tds_applicable=True,
        )

    def _period(self, start=None, end=None):
        return type('P', (), {
            'start_date': start or date(2025, 7, 1),
            'end_date': end or date(2025, 7, 31),
        })()

    def _rows(self, amount):
        return [{
            'component_code': 'BASIC',
            'amount': Decimal(amount),
            'tds_applicable': True,
            'pf_applicable': True,
            'esi_applicable': False,
            'pt_applicable': True,
        }]


class TDSRuleSeedTests(TDSComplianceTestMixin, TestCase):
    def test_seed_old_and_new_regimes(self):
        codes = set(FinancialYearTaxRule.objects.values_list('code', flat=True))
        self.assertIn('IN-TDS-NEW-2024-25', codes)
        self.assertIn('IN-TDS-OLD-2024-25', codes)
        self.assertIn('IN-TDS-NEW-2025-26', codes)
        self.assertIn('IN-TDS-OLD-2025-26', codes)
        self.assertGreaterEqual(self.new_rule.slabs.count(), 6)
        self.assertEqual(self.new_rule.standard_deduction, Decimal('75000.00'))
        self.assertEqual(self.new_rule.rebate_limit, Decimal('1200000.00'))
        self.assertEqual(self.new_rule.cess_rate, Decimal('0.0400'))

    def test_effective_date_resolution(self):
        resolved = get_tax_rule_for_fy_regime_and_date(
            '2025-26', TaxRegime.NEW, date(2025, 7, 31),
        )
        self.assertEqual(resolved.pk, self.new_rule.pk)

    def test_fy_label(self):
        self.assertEqual(financial_year_for_date(date(2025, 7, 15)), '2025-26')
        self.assertEqual(financial_year_for_date(date(2026, 3, 31)), '2025-26')
        self.assertEqual(financial_year_for_date(date(2026, 4, 1)), '2026-27')


class TDSSlabAndRebateTests(TDSComplianceTestMixin, TestCase):
    def test_slab_boundary_nil_band(self):
        slabs = list(self.new_rule.slabs.order_by('sequence'))
        # Income exactly at end of nil band (4L for FY 2025-26 NEW)
        tax = compute_slab_tax(Decimal('400000.00'), slabs)
        self.assertEqual(tax, Decimal('0.00'))

    def test_slab_above_first_band(self):
        slabs = list(self.new_rule.slabs.order_by('sequence'))
        # 500000 → 100000 in 5% band = 5000
        tax = compute_slab_tax(Decimal('500000.00'), slabs)
        self.assertEqual(tax, Decimal('5000.00'))

    def test_rebate_applies_at_limit(self):
        rebate, after = apply_rebate(
            Decimal('10000.00'),
            taxable_income=Decimal('1200000.00'),
            rebate_limit=Decimal('1200000.00'),
        )
        self.assertEqual(rebate, Decimal('10000.00'))
        self.assertEqual(after, Decimal('0.00'))

    def test_rebate_not_above_limit(self):
        rebate, after = apply_rebate(
            Decimal('10000.00'),
            taxable_income=Decimal('1200000.01'),
            rebate_limit=Decimal('1200000.00'),
        )
        self.assertEqual(rebate, Decimal('0.00'))
        self.assertEqual(after, Decimal('10000.00'))

    def test_standard_deduction_new_regime(self):
        # July period → 9 months remaining (Jul–Mar), not full 12.
        # Monthly 100000 → annual proj 900000 - std 75000 = 825000
        result = calculate_tds_for_employee(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('100000.00'),
        )
        self.assertTrue(result.applicable)
        self.assertEqual(result.tax_regime, TaxRegime.NEW)
        self.assertEqual(result.projection.standard_deduction, Decimal('75000.00'))
        self.assertEqual(result.months_left, 9)
        self.assertEqual(result.taxable_salary, Decimal('825000.00'))
        # Within rebate (≤ 12L) → annual tax 0, monthly 0
        self.assertEqual(result.annual_tax, Decimal('0.00'))
        self.assertEqual(result.monthly_tds, Decimal('0.00'))

    def test_old_regime_with_declaration_deductions(self):
        EmployeeTaxProfile.objects.filter(employee=self.employee).update(
            default_tax_regime=TaxRegime.OLD,
        )
        TaxDeclaration.objects.create(
            employee=self.employee,
            financial_year='2025-26',
            regime=TaxRegime.OLD,
            section_80c=Decimal('150000.00'),
            housing_loan=Decimal('200000.00'),
            status=TaxDeclarationStatus.APPROVED,
        )
        result = calculate_tds_for_employee(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('100000.00'),
        )
        self.assertEqual(result.tax_regime, TaxRegime.OLD)
        self.assertEqual(result.projection.standard_deduction, Decimal('50000.00'))
        self.assertEqual(result.projection.old_regime_deductions, Decimal('350000.00'))
        # Jul proj 900000 - 50k - 350k = 500000 taxable
        self.assertEqual(result.taxable_salary, Decimal('500000.00'))


class TDSPreviousEmployerAndMidYearTests(TDSComplianceTestMixin, TestCase):
    def test_previous_employer_income_and_tds(self):
        PreviousEmployerIncome.objects.create(
            employee=self.employee,
            financial_year='2025-26',
            employer_name='Prior Co',
            taxable_income=Decimal('500000.00'),
            tds_deducted=Decimal('20000.00'),
        )
        result = calculate_tds_for_employee(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('100000.00'),
        )
        self.assertEqual(result.projection.previous_employer_income, Decimal('500000.00'))
        self.assertEqual(result.projection.previous_employer_tds, Decimal('20000.00'))
        # Jul: 500k prev + 900k current-side = 1.4M - 75k = 1.325M > rebate → tax > 0
        self.assertGreater(result.annual_tax, Decimal('0.00'))
        self.assertEqual(result.previous_tds, Decimal('20000.00'))

    def test_mid_year_join(self):
        self.employee.date_of_joining = date(2025, 7, 15)
        self.employee.save()
        result = calculate_tds_for_employee(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('80000.00'),
        )
        self.assertTrue(result.applicable)
        # Still projects remaining months including current from current salary
        self.assertEqual(result.months_left, 9)  # Jul–Mar = 9


class TDSRegimeChangeTests(TDSComplianceTestMixin, TestCase):
    def test_regime_change_before_deadline(self):
        """Documented rule: change allowed until 31 July of the FY."""
        deadline = regime_change_deadline('2025-26')
        self.assertEqual(deadline, date(2025, 7, 31))
        TaxDeclaration.objects.create(
            employee=self.employee,
            financial_year='2025-26',
            regime=TaxRegime.OLD,
            status=TaxDeclarationStatus.SUBMITTED,
        )
        regime, source = resolve_regime(
            self.employee, financial_year='2025-26', as_of=date(2025, 7, 15),
        )
        self.assertEqual(regime, TaxRegime.OLD)
        self.assertEqual(source, 'tax_declaration')

    def test_declaration_still_wins_after_deadline(self):
        TaxDeclaration.objects.create(
            employee=self.employee,
            financial_year='2025-26',
            regime=TaxRegime.OLD,
            status=TaxDeclarationStatus.APPROVED,
        )
        regime, source = resolve_regime(
            self.employee, financial_year='2025-26', as_of=date(2025, 8, 15),
        )
        self.assertEqual(regime, TaxRegime.OLD)
        self.assertEqual(source, 'tax_declaration')


class TDSZeroAndSurchargeCessTests(TDSComplianceTestMixin, TestCase):
    def test_zero_taxable(self):
        result = calculate_tds_for_employee(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('0.00'),
        )
        self.assertEqual(result.taxable_salary, Decimal('0.00'))
        self.assertEqual(result.monthly_tds, Decimal('0.00'))
        self.assertEqual(result.cess, Decimal('0.00'))

    def test_cess_rate_from_rule(self):
        # High salary to produce non-zero tax after rebate band
        result = calculate_tds_for_employee(
            employee=self.employee,
            period=self._period(),
            earning_rows=self._rows('250000.00'),
        )
        self.assertGreater(result.annual_tax, Decimal('0.00'))
        if result.tax_after_rebate + result.surcharge > 0:
            expected_cess = (
                (result.tax_after_rebate + result.surcharge) * Decimal('0.04')
            ).quantize(Decimal('0.01'))
            self.assertEqual(result.cess, expected_cess)

    def test_surcharge_band(self):
        surcharge = apply_surcharge(
            Decimal('100000.00'),
            Decimal('6000000.00'),
            self.new_rule.surcharge_rules,
        )
        self.assertEqual(surcharge, Decimal('10000.00'))  # 10%


class TDSPayrollIntegrationTests(TDSComplianceTestMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.period = create_period(
            company=self.company,
            month=7,
            year=2025,
            start_date=date(2025, 7, 1),
            end_date=date(2025, 7, 31),
            user=self.user,
        )
        self.run = create_run(period=self.period, user=self.user)
        seed_standard_components(self.company)
        structure = SalaryStructure.objects.create(
            company=self.company,
            code='TDS94',
            name='TDS Structure',
        )
        basic = SalaryComponent.objects.get(company=self.company, component_code='BASIC')
        SalaryStructureLine.objects.create(
            structure=structure,
            component=basic,
            calculation_type=CalculationType.PERCENTAGE,
            percent=Decimal('100.0000'),
            display_order=10,
        )
        EmployeeSalaryAssignment.objects.create(
            employee=self.employee,
            salary_structure=structure,
            effective_from=date(2025, 4, 1),
            gross_salary=Decimal('100000.00'),
            ctc=Decimal('1200000.00'),
            created_by=self.user,
            updated_by=self.user,
        )

    def test_payroll_recalc_creates_tds_result(self):
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        self.assertIsNotNone(self.run.tds_rule_set_id)
        result = PayrollResult.objects.get(run=self.run, employee=self.employee)
        self.assertTrue(hasattr(result, 'tds_result'))
        self.assertIsInstance(result.tds_result, PayrollTDSResult)
        tds_comp = result.components.filter(component_code='STAT_TDS').first()
        self.assertIsNotNone(tds_comp)
        self.assertEqual(tds_comp.amount, result.tds_result.monthly_tds)

        # Recalc unlocked run
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        self.assertEqual(PayrollTDSResult.objects.filter(payroll_result__run=self.run).count(), 1)

    def test_locked_run_rejects_tds_clear(self):
        calculate_run(self.run, user=self.user)
        self.run.status = PayrollRunStatus.LOCKED
        self.run.save(update_fields=['status'])
        with self.assertRaises(LockedRunMutationError):
            clear_run_results(self.run)

    def test_report_reconciliation(self):
        # Bump gross so TDS is non-zero after rebate band
        assignment = self.employee.salary_assignments.first()
        assignment.gross_salary = Decimal('250000.00')
        assignment.save(update_fields=['gross_salary', 'updated_at'])
        calculate_run(self.run, user=self.user)
        resp = export_tds_register(self.run)
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertGreaterEqual(ws.max_row, 2)
        snap_tds = PayrollTDSResult.objects.get(payroll_result__run=self.run).monthly_tds
        # Column H = Monthly TDS
        self.assertEqual(Decimal(str(ws.cell(2, 8).value)), snap_tds)

        summary = export_tds_monthly_summary(self.run)
        self.assertEqual(summary.status_code, 200)


class TDSUITests(TDSComplianceTestMixin, TestCase):
    def test_hub_and_tax_pages(self):
        self.client.login(username='tdsadmin', password='TestPassword123!')
        for name in (
            'compliance:hub',
            'compliance:tax_profiles',
            'compliance:tax_declarations',
            'compliance:investment_proofs',
            'compliance:previous_employer',
        ):
            resp = self.client.get(reverse(name))
            self.assertEqual(resp.status_code, 200, msg=name)

    def test_pan_validation(self):
        profile = EmployeeTaxProfile.objects.get(employee=self.employee)
        profile.pan_number = 'BADPAN'
        with self.assertRaises(ValidationError):
            profile.full_clean()
