from datetime import datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook

from apps.employee.models import Employee

from .models import Attendance, AttendanceStatus
from .services import assert_period_editable, calculate_attendance_metrics, resolve_shift_for_employee

IMPORT_HEADERS = ['Employee Code', 'Date', 'In', 'Out', 'Status', 'OT']

STATUS_ALIASES = {
    'P': AttendanceStatus.PRESENT,
    'PRESENT': AttendanceStatus.PRESENT,
    'A': AttendanceStatus.ABSENT,
    'ABSENT': AttendanceStatus.ABSENT,
    'H': AttendanceStatus.HOLIDAY,
    'HOLIDAY': AttendanceStatus.HOLIDAY,
    'WO': AttendanceStatus.WEEKLY_OFF,
    'WEEKLY OFF': AttendanceStatus.WEEKLY_OFF,
    'WEEKLY_OFF': AttendanceStatus.WEEKLY_OFF,
    'CL': AttendanceStatus.CASUAL_LEAVE,
    'SL': AttendanceStatus.SICK_LEAVE,
    'EL': AttendanceStatus.EARNED_LEAVE,
    'LOP': AttendanceStatus.LOP,
    'HD': AttendanceStatus.HALF_DAY,
    'HALF DAY': AttendanceStatus.HALF_DAY,
    'HALF_DAY': AttendanceStatus.HALF_DAY,
    'OD': AttendanceStatus.ON_DUTY,
    'LEAVE': AttendanceStatus.CASUAL_LEAVE,
}


def _workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_attendance_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Attendance'
    sheet.append(IMPORT_HEADERS)
    sheet.append(['EMP0001', '2026-07-01', '09:00', '18:00', 'P', '0'])
    return _workbook_response(workbook, 'attendance_import_template.xlsx')


def export_import_errors(errors, filename='attendance_import_errors.xlsx'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Errors'
    sheet.append(['Row', 'Error'])
    for error in errors:
        if ': ' in error:
            row_part, message = error.split(': ', 1)
            sheet.append([row_part.replace('Row ', ''), message])
        else:
            sheet.append(['', error])
    return _workbook_response(workbook, filename)


def _cell(row, index):
    if index >= len(row):
        return ''
    value = row[index]
    if value is None:
        return ''
    return value


def _parse_date(value):
    if value is None or value == '':
        raise ValueError('Date is required')
    if hasattr(value, 'date') and callable(value.date) and not isinstance(value, str):
        try:
            return value.date()
        except Exception:
            pass
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Invalid date: {value}')


def _parse_time(value):
    if value is None or value == '':
        return None
    if isinstance(value, time):
        return value
    if hasattr(value, 'hour') and hasattr(value, 'minute') and not isinstance(value, str):
        return time(value.hour, value.minute, getattr(value, 'second', 0))
    text = str(value).strip()
    for fmt in ('%H:%M:%S', '%H:%M', '%I:%M %p'):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise ValueError(f'Invalid time: {value}')


def _parse_status(value):
    if value is None or value == '':
        return AttendanceStatus.PRESENT
    key = str(value).strip().upper()
    if key in STATUS_ALIASES:
        return STATUS_ALIASES[key]
    raise ValueError(f'Invalid status: {value}')


def _parse_ot(value):
    if value is None or value == '':
        return Decimal('0.00')
    try:
        return Decimal(str(value)).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f'Invalid OT hours: {value}') from exc


def import_attendance(company, uploaded_file, user):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    created = 0
    updated = 0
    skipped = 0
    errors = []
    seen = set()

    employees = {
        emp.employee_code.strip().upper(): emp
        for emp in Employee.objects.filter(company=company, is_active=True)
        if emp.employee_code
    }

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0, 0, 0, ['Row 1: Spreadsheet is empty']

    for index, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        try:
            code = str(_cell(row, 0)).strip().upper()
            if not code:
                raise ValueError('Employee Code is required')
            employee = employees.get(code)
            if employee is None:
                raise ValueError(f'Employee not found for company: {code}')

            attendance_date = _parse_date(_cell(row, 1))
            check_in = _parse_time(_cell(row, 2))
            check_out = _parse_time(_cell(row, 3))
            status = _parse_status(_cell(row, 4))
            overtime = _parse_ot(_cell(row, 5))

            key = (employee.pk, attendance_date.isoformat())
            if key in seen:
                raise ValueError('Duplicate row for same employee and date in file')
            seen.add(key)

            assert_period_editable(company, attendance_date, user)

            with transaction.atomic():
                existing = Attendance.objects.filter(
                    employee=employee,
                    attendance_date=attendance_date,
                ).first()
                if existing:
                    attendance = existing
                    updated += 1
                else:
                    attendance = Attendance(employee=employee, attendance_date=attendance_date)
                    created += 1
                    attendance.created_by = user

                attendance.status = status
                attendance.check_in = check_in
                attendance.check_out = check_out
                attendance.overtime_hours = overtime
                attendance.shift = resolve_shift_for_employee(employee, attendance_date)
                attendance.updated_by = user
                calculate_attendance_metrics(attendance)
                attendance.full_clean()
                attendance.save()
        except (ValidationError, ValueError) as exc:
            if isinstance(exc, ValidationError):
                if hasattr(exc, 'messages'):
                    message = '; '.join(exc.messages)
                else:
                    message = str(exc)
            else:
                message = str(exc)
            errors.append(f'Row {index}: {message}')
            skipped += 1

    return created, updated, skipped, errors


def export_attendance(queryset, filename='attendance_export.xlsx'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Attendance'
    sheet.append([
        'Employee Code',
        'Employee Name',
        'Company',
        'Date',
        'Status',
        'Shift',
        'In',
        'Out',
        'Worked Hours',
        'OT',
        'Late Minutes',
        'Early Exit',
        'Approved',
        'Remarks',
    ])
    for row in queryset:
        sheet.append([
            row.employee.employee_code,
            row.employee.full_name,
            row.employee.company.company_name,
            row.attendance_date.isoformat(),
            row.status,
            row.shift.shift_code if row.shift else '',
            row.check_in.strftime('%H:%M') if row.check_in else '',
            row.check_out.strftime('%H:%M') if row.check_out else '',
            float(row.worked_hours or 0),
            float(row.overtime_hours or 0),
            row.late_minutes,
            row.early_exit_minutes,
            'Yes' if row.approved else 'No',
            row.remarks,
        ])
    return _workbook_response(workbook, filename)
