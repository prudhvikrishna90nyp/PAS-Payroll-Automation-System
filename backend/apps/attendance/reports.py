from collections import defaultdict
from io import BytesIO

from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from openpyxl import Workbook

from .filters import attendance_list_queryset
from .models import (
    Attendance,
    AttendanceMonthlySummary,
    AttendanceStatus,
    Holiday,
    WeeklyOff,
)


def _workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _append_header(sheet, headers):
    sheet.append(headers)


def daily_register_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Daily Register'
    _append_header(
        sheet,
        ['Date', 'Employee Code', 'Employee Name', 'Company', 'Status', 'In', 'Out', 'Worked', 'OT', 'Late'],
    )
    for row in attendance_list_queryset(params).order_by('attendance_date', 'employee__employee_code'):
        sheet.append([
            row.attendance_date.isoformat(),
            row.employee.employee_code,
            row.employee.full_name,
            row.employee.company.company_name,
            row.get_status_display(),
            row.check_in.strftime('%H:%M') if row.check_in else '',
            row.check_out.strftime('%H:%M') if row.check_out else '',
            float(row.worked_hours or 0),
            float(row.overtime_hours or 0),
            row.late_minutes,
        ])
    return _workbook_response(workbook, 'attendance_daily_register.xlsx')


def monthly_register_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Monthly Register'
    _append_header(
        sheet,
        [
            'Employee Code',
            'Employee Name',
            'Present',
            'Absent',
            'Paid Leave',
            'WO',
            'Holiday',
            'Half Days',
            'OT Hours',
            'Late Count',
            'LOP',
        ],
    )
    qs = AttendanceMonthlySummary.objects.select_related('employee', 'period', 'employee__company')
    company = (params.get('company') or '').strip()
    period = (params.get('period') or '').strip()
    if company:
        qs = qs.filter(employee__company_id=company)
    if period:
        qs = qs.filter(period_id=period)
    for row in qs.order_by('employee__employee_code'):
        sheet.append([
            row.employee.employee_code,
            row.employee.full_name,
            float(row.present_days),
            float(row.absent_days),
            float(row.paid_leave_days),
            float(row.weekly_off_days),
            float(row.holiday_days),
            float(row.half_days),
            float(row.overtime_hours),
            row.late_count,
            float(row.lop_days),
        ])
    return _workbook_response(workbook, 'attendance_monthly_register.xlsx')


def late_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Late Coming'
    _append_header(sheet, ['Date', 'Employee Code', 'Employee Name', 'Late Minutes', 'In Time', 'Shift'])
    qs = attendance_list_queryset(params).filter(late_minutes__gt=0)
    for row in qs.order_by('-late_minutes', 'attendance_date'):
        sheet.append([
            row.attendance_date.isoformat(),
            row.employee.employee_code,
            row.employee.full_name,
            row.late_minutes,
            row.check_in.strftime('%H:%M') if row.check_in else '',
            row.shift.shift_code if row.shift else '',
        ])
    return _workbook_response(workbook, 'attendance_late_report.xlsx')


def overtime_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Overtime'
    _append_header(sheet, ['Date', 'Employee Code', 'Employee Name', 'OT Hours', 'Worked Hours'])
    qs = attendance_list_queryset(params).filter(overtime_hours__gt=0)
    for row in qs.order_by('-overtime_hours', 'attendance_date'):
        sheet.append([
            row.attendance_date.isoformat(),
            row.employee.employee_code,
            row.employee.full_name,
            float(row.overtime_hours),
            float(row.worked_hours or 0),
        ])
    return _workbook_response(workbook, 'attendance_ot_report.xlsx')


def absentee_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Absentee'
    _append_header(sheet, ['Date', 'Employee Code', 'Employee Name', 'Status', 'Remarks'])
    qs = attendance_list_queryset(params).filter(
        status__in=[AttendanceStatus.ABSENT, AttendanceStatus.LOP]
    )
    for row in qs.order_by('attendance_date', 'employee__employee_code'):
        sheet.append([
            row.attendance_date.isoformat(),
            row.employee.employee_code,
            row.employee.full_name,
            row.get_status_display(),
            row.remarks,
        ])
    return _workbook_response(workbook, 'attendance_absentee_report.xlsx')


def weekly_off_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Weekly Off'
    _append_header(sheet, ['Employee Code', 'Employee Name', 'Weekday', 'Effective From', 'Effective To'])
    qs = WeeklyOff.objects.select_related('employee', 'employee__company').filter(is_active=True)
    company = (params.get('company') or '').strip()
    if company:
        qs = qs.filter(employee__company_id=company)
    for row in qs.order_by('employee__employee_code', 'weekday'):
        sheet.append([
            row.employee.employee_code,
            row.employee.full_name,
            row.get_weekday_display(),
            row.effective_from.isoformat(),
            row.effective_to.isoformat() if row.effective_to else '',
        ])
    return _workbook_response(workbook, 'attendance_weekly_off_report.xlsx')


def holiday_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Holidays'
    _append_header(sheet, ['Date', 'Holiday', 'Type', 'Company'])
    qs = Holiday.objects.select_related('company').filter(is_active=True)
    company = (params.get('company') or '').strip()
    year = (params.get('year') or '').strip()
    if company:
        qs = qs.filter(company_id=company)
    if year:
        qs = qs.filter(holiday_date__year=year)
    for row in qs.order_by('holiday_date'):
        sheet.append([
            row.holiday_date.isoformat(),
            row.holiday_name,
            row.get_holiday_type_display(),
            row.company.company_name,
        ])
    return _workbook_response(workbook, 'attendance_holiday_report.xlsx')


def summary_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Summary'
    _append_header(
        sheet,
        ['Status', 'Count', 'Total OT Hours', 'Late Records'],
    )
    qs = attendance_list_queryset(params)
    grouped = qs.values('status').annotate(
        count=Count('id'),
        ot=Sum('overtime_hours'),
        late_rows=Count('id', filter=Q(late_minutes__gt=0)),
    ).order_by('status')
    status_labels = dict(AttendanceStatus.choices)
    totals = defaultdict(int)
    for row in grouped:
        sheet.append([
            status_labels.get(row['status'], row['status']),
            row['count'],
            float(row['ot'] or 0),
            row['late_rows'],
        ])
        totals['count'] += row['count']
    sheet.append([])
    sheet.append(['Total records', totals['count']])
    return _workbook_response(workbook, 'attendance_summary_report.xlsx')


REPORT_BUILDERS = {
    'daily': daily_register_report,
    'monthly': monthly_register_report,
    'late': late_report,
    'ot': overtime_report,
    'absentee': absentee_report,
    'weekly_off': weekly_off_report,
    'holiday': holiday_report,
    'summary': summary_report,
}
