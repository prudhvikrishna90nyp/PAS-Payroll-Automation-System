from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook

from .filters import assignment_list_queryset, component_list_queryset, structure_list_queryset
from .models import EmployeeSalaryAssignment
from .services.salary_calculator import calculate_assignment_components


def _workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def component_register_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Component Register'
    sheet.append([
        'Company', 'Code', 'Name', 'Type', 'Calc Type', 'Formula',
        'Taxable', 'PF', 'ESI', 'In CTC', 'In Gross', 'Order', 'Active',
    ])
    for row in component_list_queryset(params):
        sheet.append([
            row.company.company_name,
            row.component_code,
            row.component_name,
            row.get_component_type_display(),
            row.get_calculation_type_display(),
            row.formula,
            'Yes' if row.taxable else 'No',
            'Yes' if row.pf_applicable else 'No',
            'Yes' if row.esi_applicable else 'No',
            'Yes' if row.include_in_ctc else 'No',
            'Yes' if row.include_in_gross else 'No',
            row.display_order,
            'Yes' if row.is_active else 'No',
        ])
    return _workbook_response(workbook, 'component_register.xlsx')


def structure_register_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Structure Register'
    sheet.append([
        'Company', 'Structure Code', 'Structure Name', 'Component Code',
        'Component Name', 'Calc Type', 'Value', 'Percent', 'Formula', 'Order', 'Active',
    ])
    for structure in structure_list_queryset(params).prefetch_related('lines__component'):
        lines = list(structure.lines.all())
        if not lines:
            sheet.append([
                structure.company.company_name,
                structure.code,
                structure.name,
                '', '', '', '', '', '', '',
                'Yes' if structure.is_active else 'No',
            ])
            continue
        for line in lines:
            sheet.append([
                structure.company.company_name,
                structure.code,
                structure.name,
                line.component.component_code,
                line.component.component_name,
                line.get_calculation_type_display() if line.calculation_type else line.component.get_calculation_type_display(),
                float(line.value) if line.value is not None else '',
                float(line.percent) if line.percent is not None else '',
                line.formula_override or line.component.formula,
                line.display_order,
                'Yes' if structure.is_active else 'No',
            ])
    return _workbook_response(workbook, 'structure_register.xlsx')


def employee_salary_register_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Employee Salary Register'
    sheet.append([
        'Employee Code', 'Employee Name', 'Company', 'Structure',
        'Effective From', 'Effective To', 'Gross', 'CTC', 'Current',
    ])
    for row in assignment_list_queryset(params):
        sheet.append([
            row.employee.employee_code,
            row.employee.full_name,
            row.employee.company.company_name,
            row.salary_structure.name,
            row.effective_from.isoformat(),
            row.effective_to.isoformat() if row.effective_to else '',
            float(row.gross_salary),
            float(row.ctc),
            'Yes' if row.is_current else 'No',
        ])
    return _workbook_response(workbook, 'employee_salary_register.xlsx')


def salary_revision_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Salary Revision'
    sheet.append([
        'Employee Code', 'Employee Name', 'Company', 'Structure',
        'Effective From', 'Effective To', 'Gross', 'CTC', 'Remarks',
    ])
    qs = assignment_list_queryset(params).order_by(
        'employee__employee_code', 'effective_from'
    )
    for row in qs:
        sheet.append([
            row.employee.employee_code,
            row.employee.full_name,
            row.employee.company.company_name,
            f'{row.salary_structure.code} — {row.salary_structure.name}',
            row.effective_from.isoformat(),
            row.effective_to.isoformat() if row.effective_to else 'Current',
            float(row.gross_salary),
            float(row.ctc),
            row.remarks,
        ])
    return _workbook_response(workbook, 'salary_revision_report.xlsx')


def ctc_register_report(params):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'CTC Register'
    sheet.append([
        'Employee Code', 'Employee Name', 'Company', 'Structure',
        'Gross (Monthly)', 'CTC (Annual)', 'CTC (Monthly calc)', 'Basic',
        'Effective From',
    ])
    params = params.copy() if hasattr(params, 'copy') else dict(params)
    if hasattr(params, 'setlist'):
        params['current'] = '1'
    else:
        params = {**params, 'current': '1'}
    for row in assignment_list_queryset(params):
        basic = ''
        ctc_monthly = ''
        try:
            result = calculate_assignment_components(row)
            basic_amt = result.amounts.get('BASIC')
            if basic_amt is None:
                for code, amount in result.amounts.items():
                    if code.upper() == 'BASIC':
                        basic_amt = amount
                        break
            basic = float(basic_amt) if basic_amt is not None else ''
            ctc_monthly = float(result.ctc_monthly)
        except Exception:
            pass
        annual = float(row.ctc) if row.ctc else (ctc_monthly * 12 if ctc_monthly != '' else '')
        sheet.append([
            row.employee.employee_code,
            row.employee.full_name,
            row.employee.company.company_name,
            row.salary_structure.name,
            float(row.gross_salary),
            annual if annual != '' else '',
            ctc_monthly,
            basic,
            row.effective_from.isoformat(),
        ])
    return _workbook_response(workbook, 'ctc_register.xlsx')


REPORT_BUILDERS = {
    'component_register': component_register_report,
    'structure_register': structure_register_report,
    'employee_salary_register': employee_salary_register_report,
    'salary_revision': salary_revision_report,
    'ctc_register': ctc_register_report,
}
