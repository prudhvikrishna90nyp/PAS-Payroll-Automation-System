from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.core.exceptions import PermissionDenied, ValidationError
from django.db.models import Sum
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.clients.models import Client
from apps.company.models import Branch, Company, Department
from apps.employee.models import Employee
from apps.compliance.services.pf_rules import seed_default_pf_rule_set

from .models import (
    CalculationType,
    ComponentType,
    EmployeeSalaryAssignment,
    PayPeriod,
    PayrollAuditLog,
    PayrollPeriod,
    PayrollPeriodStatus,
    PayrollResult,
    PayrollResultComponent,
    PayrollRun,
    PayrollRunStatus,
    RoundingRule,
    SalaryComponent,
    SalaryStructure,
    SalaryStructureLine,
)
from .permissions import seed_role_groups
from .seed import seed_standard_components
from .services.approval import approve_run, mark_reviewed
from .services.calculator import calculate_employee, compute_payable_days, compute_proration_factor
from .services.exceptions import (
    InvalidTransitionError,
    LockedRunError,
    RunNotCalculableError,
    RunNotReadyError,
)
from .services.formula_engine import (
    FormulaError,
    detect_circular_references,
    evaluate_formula,
    extract_references,
)
from .services.locking import lock_run, reopen_run
from .services.payroll_engine import (
    calculate_run,
    close_period,
    create_period,
    create_run,
    open_period,
)
from .services.payslip_generator import calculate_payslip_amounts, generate_payslip
from .services.payslip_data import build_payslip_dataset
from .services.export_service import export_payroll_register
from .services.report_queries import (
    ReportFilters,
    aggregate_result_totals,
    branch_summary,
    companies_visible_to_user,
    department_summary,
    reconcile_earnings_deductions,
    results_queryset,
)
from .services.salary_calculator import apply_rounding, calculate_structure_components
from .services.validation import (
    SalaryValidationError,
    validate_component,
    validate_payroll_period,
    validate_structure,
)


def _perm(codename):
    return Permission.objects.get(codename=codename, content_type__app_label='payroll')


class PayrollTestMixin:
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='payadmin',
            password='TestPassword123!',
        )
        for codename in (
            'view_salarycomponent',
            'add_salarycomponent',
            'change_salarycomponent',
            'delete_salarycomponent',
            'export_salarycomponent',
            'view_salarystructure',
            'add_salarystructure',
            'change_salarystructure',
            'delete_salarystructure',
            'export_salarystructure',
            'view_salarystructureline',
            'add_salarystructureline',
            'change_salarystructureline',
            'delete_salarystructureline',
            'view_employeesalaryassignment',
            'add_employeesalaryassignment',
            'change_employeesalaryassignment',
            'delete_employeesalaryassignment',
            'export_employeesalaryassignment',
            'view_payslip',
            'add_payslip',
            'change_payslip',
            'view_payrollperiod',
            'add_payrollperiod',
            'change_payrollperiod',
            'view_payrollrun',
            'add_payrollrun',
            'change_payrollrun',
            'view_payrollresult',
        ):
            try:
                self.user.user_permissions.add(_perm(codename))
            except Permission.DoesNotExist:
                pass

        self.client_record = Client.objects.create(
            client_code='PAYCLI',
            client_name='Payroll Client',
            mobile='9876543210',
            address_line_1='Naidupet',
            city='Naidupet',
            state='Andhra Pradesh',
            pincode='524126',
        )
        self.company = Company.objects.create(
            client=self.client_record,
            company_name='Payroll Co',
        )
        self.employee = Employee.objects.create(
            company=self.company,
            employee_code='EMP7001',
            first_name='Ravi',
            last_name='Kumar',
            date_of_joining=date(2026, 1, 1),
            basic_salary=Decimal('20000.00'),
            auto_generate_code=False,
        )
        seed_default_pf_rule_set()


class FormulaEngineTests(TestCase):
    def test_basic_percentage_of_gross(self):
        result = evaluate_formula('Gross * 40 / 100', {'GROSS': Decimal('50000')})
        self.assertEqual(result, Decimal('20000.00'))

    def test_hra_of_basic(self):
        result = evaluate_formula(
            'Basic * 40 / 100',
            {'BASIC': Decimal('20000'), 'GROSS': Decimal('50000')},
        )
        self.assertEqual(result, Decimal('8000.00'))

    def test_special_balance_formula(self):
        result = evaluate_formula(
            'Gross - (Basic + HRA)',
            {
                'GROSS': Decimal('50000'),
                'BASIC': Decimal('20000'),
                'HRA': Decimal('8000'),
            },
        )
        self.assertEqual(result, Decimal('22000.00'))

    def test_assignment_style_formula(self):
        result = evaluate_formula('Basic = Gross * 40 / 100', {'GROSS': Decimal('10000')})
        self.assertEqual(result, Decimal('4000.00'))

    def test_rejects_eval_and_calls(self):
        with self.assertRaises(FormulaError):
            evaluate_formula('__import__("os").system("echo hi")', {'GROSS': 1})
        with self.assertRaises(FormulaError):
            evaluate_formula('abs(Gross)', {'GROSS': 10})

    def test_division_by_zero(self):
        with self.assertRaises(FormulaError):
            evaluate_formula('Gross / 0', {'GROSS': 10})

    def test_extract_references(self):
        refs = extract_references('Gross - (Basic + HRA + Conveyance)')
        self.assertIn('GROSS', refs)
        self.assertIn('BASIC', refs)
        self.assertIn('HRA', refs)
        self.assertIn('CONVEYANCE', refs)

    def test_circular_detection(self):
        cycles = detect_circular_references({
            'A': {'B'},
            'B': {'C'},
            'C': {'A'},
        })
        self.assertTrue(cycles)


class SalaryCalculatorTests(PayrollTestMixin, TestCase):
    def _office_structure(self):
        seed_standard_components(self.company)
        structure = SalaryStructure.objects.create(
            company=self.company,
            code='OFFICE',
            name='Office Staff',
        )
        for code, formula, order in (
            ('BASIC', 'Gross * 40 / 100', 10),
            ('HRA', 'Basic * 40 / 100', 20),
            ('CONVEYANCE', None, 30),
            ('SPECIAL', 'Gross - (Basic + HRA + Conveyance)', 40),
        ):
            component = SalaryComponent.objects.get(company=self.company, component_code=code)
            SalaryStructureLine.objects.create(
                structure=structure,
                component=component,
                calculation_type=CalculationType.FORMULA if formula else CalculationType.FIXED,
                formula_override=formula or '',
                value=Decimal('1600.00') if code == 'CONVEYANCE' else None,
                display_order=order,
            )
        return structure

    def test_standard_breakdown(self):
        structure = self._office_structure()
        result = calculate_structure_components(structure, Decimal('50000.00'))
        self.assertEqual(result.amounts['BASIC'], Decimal('20000.00'))
        self.assertEqual(result.amounts['HRA'], Decimal('8000.00'))
        self.assertEqual(result.amounts['CONVEYANCE'], Decimal('1600.00'))
        self.assertEqual(result.amounts['SPECIAL'], Decimal('20400.00'))

    def test_circular_formula_raises(self):
        a = SalaryComponent.objects.create(
            company=self.company,
            component_code='A1',
            component_name='A',
            component_type=ComponentType.EARNING,
            calculation_type=CalculationType.FORMULA,
            formula='B1',
        )
        b = SalaryComponent.objects.create(
            company=self.company,
            component_code='B1',
            component_name='B',
            component_type=ComponentType.EARNING,
            calculation_type=CalculationType.FORMULA,
            formula='A1',
        )
        structure = SalaryStructure.objects.create(
            company=self.company, code='CIRC', name='Circular'
        )
        SalaryStructureLine.objects.create(structure=structure, component=a, display_order=1)
        SalaryStructureLine.objects.create(structure=structure, component=b, display_order=2)
        with self.assertRaises(FormulaError):
            calculate_structure_components(structure, Decimal('10000'))


class ValidationTests(PayrollTestMixin, TestCase):
    def test_duplicate_component_code(self):
        SalaryComponent.objects.create(
            company=self.company,
            component_code='BASIC',
            component_name='Basic',
            component_type=ComponentType.EARNING,
        )
        dup = SalaryComponent(
            company=self.company,
            component_code='basic',
            component_name='Basic Two',
            component_type=ComponentType.EARNING,
        )
        errors = validate_component(dup)
        self.assertTrue(any('Duplicate component code' in e for e in errors))

    def test_missing_basic_in_structure(self):
        hra = SalaryComponent.objects.create(
            company=self.company,
            component_code='HRA',
            component_name='HRA',
            component_type=ComponentType.EARNING,
            calculation_type=CalculationType.FIXED,
        )
        structure = SalaryStructure.objects.create(
            company=self.company, code='NOBASIC', name='No Basic'
        )
        SalaryStructureLine.objects.create(
            structure=structure, component=hra, value=Decimal('1000'), display_order=1
        )
        errors = validate_structure(structure)
        self.assertTrue(any('BASIC' in e for e in errors))


class SeedAndAssignmentTests(PayrollTestMixin, TestCase):
    def test_seed_standard_components_idempotent(self):
        first = seed_standard_components(self.company)
        second = seed_standard_components(self.company)
        self.assertGreaterEqual(len(first), 15)
        self.assertEqual(len(second), 0)
        self.assertTrue(
            SalaryComponent.objects.filter(company=self.company, component_code='BASIC').exists()
        )

    def test_assignment_soft_ends_previous(self):
        seed_standard_components(self.company)
        structure = SalaryStructure.objects.create(
            company=self.company, code='OFF', name='Office'
        )
        basic = SalaryComponent.objects.get(company=self.company, component_code='BASIC')
        SalaryStructureLine.objects.create(
            structure=structure,
            component=basic,
            calculation_type=CalculationType.FORMULA,
            formula_override='Gross * 40 / 100',
            display_order=10,
        )
        first = EmployeeSalaryAssignment.objects.create(
            employee=self.employee,
            salary_structure=structure,
            effective_from=date(2026, 1, 1),
            gross_salary=Decimal('40000.00'),
            ctc=Decimal('480000.00'),
        )
        second = EmployeeSalaryAssignment.objects.create(
            employee=self.employee,
            salary_structure=structure,
            effective_from=date(2026, 7, 1),
            gross_salary=Decimal('50000.00'),
            ctc=Decimal('600000.00'),
        )
        first.refresh_from_db()
        self.assertEqual(first.effective_to, date(2026, 6, 30))
        self.assertIsNone(second.effective_to)
        self.employee.refresh_from_db()
        self.assertEqual(self.employee.basic_salary, Decimal('20000.00'))


class PayslipIntegrationTests(PayrollTestMixin, TestCase):
    def test_legacy_payslip_still_works(self):
        period = PayPeriod.objects.create(
            year=2026,
            month=7,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
        )
        gross, deductions, net, earnings, deduction_items = calculate_payslip_amounts(
            self.employee, period
        )
        self.assertGreater(gross, 0)
        self.assertEqual(len(earnings), 3)
        payslip = generate_payslip(self.employee, period)
        self.assertEqual(payslip.employee_id, self.employee.pk)
        self.assertTrue(payslip.items.exists())

    def test_structure_aware_payslip(self):
        seed_standard_components(self.company)
        structure = SalaryStructure.objects.create(
            company=self.company, code='OFF', name='Office'
        )
        for code, formula, value in (
            ('BASIC', 'Gross * 40 / 100', None),
            ('HRA', 'Basic * 40 / 100', None),
            ('CONVEYANCE', None, Decimal('1600')),
        ):
            component = SalaryComponent.objects.get(company=self.company, component_code=code)
            SalaryStructureLine.objects.create(
                structure=structure,
                component=component,
                calculation_type=CalculationType.FORMULA if formula else CalculationType.FIXED,
                formula_override=formula or '',
                value=value,
            )
        EmployeeSalaryAssignment.objects.create(
            employee=self.employee,
            salary_structure=structure,
            effective_from=date(2026, 1, 1),
            gross_salary=Decimal('50000.00'),
            ctc=Decimal('600000.00'),
        )
        period = PayPeriod.objects.create(
            year=2026,
            month=7,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
        )
        payslip = generate_payslip(self.employee, period)
        self.assertEqual(payslip.basic_salary, Decimal('20000.00'))
        self.assertEqual(payslip.gross_pay, Decimal('50000.00'))


class PayrollPermissionViewTests(PayrollTestMixin, TestCase):
    def test_component_list_requires_login(self):
        url = reverse('payroll:component_list')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)

    def test_component_list_ok(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('payroll:component_list'))
        self.assertEqual(response.status_code, 200)

    def test_structure_list_ok(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('payroll:structure_list'))
        self.assertEqual(response.status_code, 200)

    def test_assignment_list_ok(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('payroll:assignment_list'))
        self.assertEqual(response.status_code, 200)

    def test_seed_components_view(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('payroll:component_seed'),
            {'company': self.company.pk},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            SalaryComponent.objects.filter(company=self.company, component_code='EE_PF').exists()
        )

    def test_component_create(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('payroll:component_add'),
            {
                'company': self.company.pk,
                'component_code': 'TEST',
                'component_name': 'Test Comp',
                'component_type': ComponentType.EARNING,
                'calculation_type': CalculationType.FIXED,
                'formula': '',
                'rounding_rule': 'nearest',
                'display_order': 50,
                'taxable': True,
                'include_in_ctc': True,
                'include_in_gross': True,
                'is_active': True,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            SalaryComponent.objects.filter(company=self.company, component_code='TEST').exists()
        )

    def test_reports_download(self):
        self.client.force_login(self.user)
        seed_standard_components(self.company)
        for report in (
            'component_register',
            'structure_register',
            'employee_salary_register',
            'salary_revision',
            'ctc_register',
        ):
            response = self.client.get(reverse('payroll:report_download', args=[report]))
            self.assertEqual(response.status_code, 200, report)
            self.assertIn(
                'spreadsheetml',
                response['Content-Type'],
            )

    def test_payslip_list_namespaced(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('payroll:payslip_list'))
        self.assertEqual(response.status_code, 200)

    def test_seed_role_groups(self):
        names = seed_role_groups()
        self.assertIn('Payroll', names)


class NegativeSalaryValidationTests(PayrollTestMixin, TestCase):
    def test_negative_gross_rejected_by_calculator(self):
        structure = SalaryStructure.objects.create(
            company=self.company, code='X', name='X'
        )
        with self.assertRaises(FormulaError):
            calculate_structure_components(structure, Decimal('-1'))


class PayrollPeriodFoundationTests(PayrollTestMixin, TestCase):
    def test_period_unique_company_month_year(self):
        create_period(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            user=self.user,
        )
        with self.assertRaises(SalaryValidationError):
            create_period(
                company=self.company,
                month=7,
                year=2026,
                start_date=date(2026, 7, 1),
                end_date=date(2026, 7, 31),
                user=self.user,
            )

    def test_period_overlap_prevention(self):
        create_period(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            user=self.user,
        )
        period = PayrollPeriod(
            company=self.company,
            month=8,
            year=2026,
            start_date=date(2026, 7, 15),
            end_date=date(2026, 8, 14),
        )
        errors = validate_payroll_period(period)
        self.assertTrue(any('overlaps' in e.lower() for e in errors))

    def test_open_close_period(self):
        period = create_period(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            user=self.user,
        )
        self.assertEqual(period.status, PayrollPeriodStatus.OPEN)
        close_period(period, user=self.user)
        period.refresh_from_db()
        self.assertEqual(period.status, PayrollPeriodStatus.CLOSED)
        open_period(period, user=self.user)
        period.refresh_from_db()
        self.assertEqual(period.status, PayrollPeriodStatus.OPEN)

    def test_audit_log_on_period_close(self):
        period = create_period(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            user=self.user,
        )
        close_period(period, user=self.user)
        self.assertTrue(
            PayrollAuditLog.objects.filter(period=period, action='period_close').exists()
        )


class PayrollRunFoundationTests(PayrollTestMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.period = create_period(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            user=self.user,
        )

    def test_create_draft_run(self):
        run = create_run(period=self.period, user=self.user, notes='First draft')
        self.assertEqual(run.status, PayrollRunStatus.DRAFT)
        self.assertEqual(run.run_number, 1)
        self.assertEqual(run.company_id, self.company.pk)

    def test_audit_log_on_run_create(self):
        run = create_run(period=self.period, user=self.user)
        self.assertTrue(
            PayrollAuditLog.objects.filter(run=run, action='run_create').exists()
        )

    def test_cannot_create_run_on_closed_period(self):
        close_period(self.period, user=self.user)
        with self.assertRaises(SalaryValidationError):
            create_run(period=self.period, user=self.user)

    def test_run_number_increments(self):
        create_run(period=self.period, user=self.user)
        run2 = create_run(period=self.period, user=self.user)
        self.assertEqual(run2.run_number, 2)


class PayrollEngineAuthViewTests(PayrollTestMixin, TestCase):
    def test_period_list_requires_login(self):
        response = self.client.get(reverse('payroll:period_list'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.url)

    def test_run_list_requires_login(self):
        response = self.client.get(reverse('payroll:run_list'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.url)

    def test_period_list_ok(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('payroll:period_list'))
        self.assertEqual(response.status_code, 200)

    def test_run_create_view(self):
        period = create_period(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            user=self.user,
        )
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('payroll:run_add'),
            {'period': period.pk, 'notes': 'UI create'},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            PayrollRun.objects.filter(period=period, status=PayrollRunStatus.DRAFT).exists()
        )


class PayrollCalculationEngineTests(PayrollTestMixin, TestCase):
    """Sprint 8.2 — calculate_run pipeline and proration."""

    def setUp(self):
        super().setUp()
        self.period = create_period(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            user=self.user,
        )
        self.run = create_run(period=self.period, user=self.user)
        self.structure = self._build_structure()
        self.assignment = EmployeeSalaryAssignment.objects.create(
            employee=self.employee,
            salary_structure=self.structure,
            effective_from=date(2026, 1, 1),
            gross_salary=Decimal('50000.00'),
            ctc=Decimal('600000.00'),
            created_by=self.user,
            updated_by=self.user,
        )

    def _build_structure(self):
        seed_standard_components(self.company)
        structure = SalaryStructure.objects.create(
            company=self.company,
            code='CALC82',
            name='Calc Engine Structure',
        )
        basic = SalaryComponent.objects.get(company=self.company, component_code='BASIC')
        hra = SalaryComponent.objects.get(company=self.company, component_code='HRA')
        conv = SalaryComponent.objects.get(company=self.company, component_code='CONVEYANCE')
        SalaryStructureLine.objects.create(
            structure=structure,
            component=basic,
            calculation_type=CalculationType.PERCENTAGE,
            percent=Decimal('40.0000'),
            display_order=10,
        )
        SalaryStructureLine.objects.create(
            structure=structure,
            component=hra,
            calculation_type=CalculationType.FORMULA,
            formula_override='Basic * 40 / 100',
            display_order=20,
        )
        SalaryStructureLine.objects.create(
            structure=structure,
            component=conv,
            calculation_type=CalculationType.FIXED,
            value=Decimal('1600.00'),
            display_order=30,
        )
        return structure

    def _att_period(self):
        from apps.attendance.models import AttendancePeriod, PeriodStatus

        return AttendancePeriod.objects.create(
            company=self.company,
            month=7,
            year=2026,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            status=PeriodStatus.LOCKED,
        )

    def test_full_month_employee(self):
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        result = PayrollResult.objects.get(run=self.run, employee=self.employee)
        # 40% basic + 40% HRA of basic + 1600 conveyance = 20000 + 8000 + 1600
        self.assertEqual(result.total_earnings, Decimal('29600.00'))
        self.assertEqual(result.gross, Decimal('29600.00'))
        self.assertEqual(result.net_salary, result.gross - result.total_deductions)
        codes = set(
            PayrollResultComponent.objects.filter(result=result).values_list(
                'component_code', flat=True
            )
        )
        self.assertIn('BASIC', codes)
        self.assertIn('STAT_PF', codes)
        pf = PayrollResultComponent.objects.get(result=result, component_code='STAT_PF')
        # BASIC 20000 capped at 15000 × 12% = 1800 (Sprint 9.1 EPF)
        self.assertEqual(pf.amount, Decimal('1800.00'))
        self.assertEqual(result.total_deductions, Decimal('1800.00'))
        self.assertEqual(self.run.pf_rule_set_id is not None, True)
        self.assertTrue(hasattr(result, 'pf_result'))
        self.assertEqual(result.pf_result.employee_pf, Decimal('1800.00'))
        self.assertEqual(result.pf_result.eps, Decimal('1249.50'))  # 15000 * 8.33%

    def test_mid_month_joining_proration(self):
        self.employee.date_of_joining = date(2026, 7, 16)
        self.employee.save(update_fields=['date_of_joining'])
        calc = calculate_employee(employee=self.employee, period=self.period)
        # Jul 16–31 = 16 eligible / 31 calendar
        self.assertEqual(calc.eligible_days, Decimal('16'))
        self.assertEqual(calc.calendar_days, Decimal('31'))
        expected_factor = compute_proration_factor(Decimal('16'), Decimal('31'))
        self.assertEqual(calc.proration_factor, expected_factor)
        full_earnings = Decimal('29600.00')
        expected = (full_earnings * expected_factor).quantize(Decimal('0.01'))
        self.assertEqual(calc.total_earnings, expected)

    def test_lop_proration(self):
        from apps.attendance.models import AttendanceMonthlySummary

        att_period = self._att_period()
        # 31 eligible; present+leave+WO+holiday = 28 → payable 28 (3 LOP implied)
        AttendanceMonthlySummary.objects.create(
            employee=self.employee,
            period=att_period,
            present_days=Decimal('20.00'),
            paid_leave_days=Decimal('2.00'),
            weekly_off_days=Decimal('4.00'),
            holiday_days=Decimal('2.00'),
            lop_days=Decimal('3.00'),
            absent_days=Decimal('3.00'),
        )
        calc = calculate_employee(employee=self.employee, period=self.period)
        self.assertEqual(calc.payable_days, Decimal('28.00'))
        factor = compute_proration_factor(Decimal('28'), Decimal('31'))
        self.assertEqual(calc.proration_factor, factor)

    def test_half_day_counts_as_half_present(self):
        from apps.attendance.models import AttendanceMonthlySummary
        from apps.payroll.services.attendance_loader import AttendanceSnapshot

        # present_days already includes half-day * 0.5 per attendance services
        snap = AttendanceSnapshot(
            present_days=Decimal('0.50'),
            half_days=Decimal('1.00'),
            source='summary',
        )
        payable = compute_payable_days(eligible_days=Decimal('31'), attendance=snap)
        self.assertEqual(payable, Decimal('0.50'))

        att_period = self._att_period()
        AttendanceMonthlySummary.objects.create(
            employee=self.employee,
            period=att_period,
            present_days=Decimal('0.50'),
            half_days=Decimal('1.00'),
            lop_days=Decimal('0.00'),
        )
        calc = calculate_employee(employee=self.employee, period=self.period)
        self.assertEqual(calc.payable_days, Decimal('0.50'))
        self.assertEqual(calc.present_days, Decimal('0.50'))

    def test_fixed_percentage_formula_components(self):
        calculate_run(self.run, user=self.user)
        result = PayrollResult.objects.get(run=self.run, employee=self.employee)
        by_code = {
            c.component_code: c.amount
            for c in result.components.all()
            if c.component_type == ComponentType.EARNING
        }
        self.assertEqual(by_code['BASIC'], Decimal('20000.00'))
        self.assertEqual(by_code['HRA'], Decimal('8000.00'))
        self.assertEqual(by_code['CONVEYANCE'], Decimal('1600.00'))

    def test_rounding_nearest_rupee(self):
        component = SalaryComponent.objects.create(
            company=self.company,
            component_code='ROUNDT',
            component_name='Rounding Test',
            component_type=ComponentType.EARNING,
            calculation_type=CalculationType.FIXED,
            rounding_rule=RoundingRule.NEAREST,
        )
        amount = apply_rounding(Decimal('100.60'), RoundingRule.NEAREST)
        self.assertEqual(amount, Decimal('101.00'))
        amount_down = apply_rounding(Decimal('100.40'), RoundingRule.NEAREST)
        self.assertEqual(amount_down, Decimal('100.00'))
        self.assertEqual(component.rounding_rule, RoundingRule.NEAREST)

    def test_missing_salary_assignment(self):
        self.assignment.soft_delete()
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.INCOMPLETE)
        self.assertTrue(self.run.calculation_errors)
        self.assertEqual(self.run.calculation_errors[0]['code'], 'missing_salary_assignment')
        self.assertFalse(
            PayrollResult.objects.filter(run=self.run, employee=self.employee).exists()
        )

    def test_invalid_formula(self):
        SalaryStructureLine.objects.filter(
            structure=self.structure, component__component_code='HRA'
        ).update(formula_override='Basic / 0')
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.INCOMPLETE)
        self.assertIn('Division by zero', self.run.calculation_errors[0]['error'])
        self.assertFalse(
            PayrollResult.objects.filter(run=self.run, employee=self.employee).exists()
        )

    def test_circular_formula(self):
        SalaryStructureLine.objects.filter(
            structure=self.structure, component__component_code='BASIC'
        ).update(
            calculation_type=CalculationType.FORMULA,
            formula_override='HRA',
            percent=None,
        )
        SalaryStructureLine.objects.filter(
            structure=self.structure, component__component_code='HRA'
        ).update(formula_override='Basic')
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.INCOMPLETE)
        self.assertTrue(
            any('circular' in e['error'].lower() for e in self.run.calculation_errors)
        )

    def test_locked_run_rejection(self):
        self.run.status = PayrollRunStatus.LOCKED
        self.run.save(update_fields=['status'])
        with self.assertRaises(LockedRunError):
            calculate_run(self.run, user=self.user)

    def test_reviewed_run_rejection(self):
        self.run.status = PayrollRunStatus.REVIEWED
        self.run.save(update_fields=['status'])
        with self.assertRaises(RunNotCalculableError):
            calculate_run(self.run, user=self.user)

    def test_recalculation_unlocked_replaces_results(self):
        calculate_run(self.run, user=self.user)
        first_id = PayrollResult.objects.get(run=self.run).pk
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        self.assertEqual(PayrollResult.objects.filter(run=self.run).count(), 1)
        self.assertNotEqual(PayrollResult.objects.get(run=self.run).pk, first_id)

    def test_duplicate_calc_prevention_for_locked(self):
        calculate_run(self.run, user=self.user)
        self.run.status = PayrollRunStatus.LOCKED
        self.run.save(update_fields=['status'])
        with self.assertRaises(LockedRunError):
            calculate_run(self.run, user=self.user)
        self.assertEqual(PayrollResult.objects.filter(run=self.run).count(), 1)

    def test_transaction_rollback_on_unexpected_failure(self):
        calculate_run(self.run, user=self.user)
        self.assertEqual(PayrollResult.objects.filter(run=self.run).count(), 1)
        from unittest.mock import patch

        with patch(
            'apps.payroll.services.payroll_engine.snapshot_employee_result',
            side_effect=RuntimeError('boom'),
        ):
            with self.assertRaises(RuntimeError):
                calculate_run(self.run, user=self.user)
        # Outer atomic rolls back: prior successful result remains (delete+fail undone)
        self.run.refresh_from_db()
        self.assertEqual(PayrollResult.objects.filter(run=self.run).count(), 1)
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)

    def test_partial_failure_keeps_successful_employee(self):
        emp2 = Employee.objects.create(
            company=self.company,
            employee_code='EMP7002',
            first_name='No',
            last_name='Assign',
            date_of_joining=date(2026, 1, 1),
            basic_salary=Decimal('10000.00'),
            auto_generate_code=False,
        )
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.INCOMPLETE)
        self.assertTrue(
            PayrollResult.objects.filter(run=self.run, employee=self.employee).exists()
        )
        self.assertFalse(
            PayrollResult.objects.filter(run=self.run, employee=emp2).exists()
        )
        self.assertEqual(len(self.run.calculation_errors), 1)
        self.assertEqual(self.run.calculation_errors[0]['employee_code'], 'EMP7002')

    def test_calculate_view_ui(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('payroll:run_calculate', args=[self.run.pk]))
        self.assertEqual(response.status_code, 302)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        detail = self.client.get(reverse('payroll:run_detail', args=[self.run.pk]))
        self.assertEqual(detail.status_code, 200)
        self.assertContains(detail, 'Totals')
        self.assertContains(detail, 'EMP7001')


class PayrollApprovalLockingTests(PayrollTestMixin, TestCase):
    """Sprint 8.3 — review / approve / lock workflow."""

    def setUp(self):
        super().setUp()
        seed_role_groups()
        self.period = create_period(
            company=self.company,
            month=8,
            year=2026,
            start_date=date(2026, 8, 1),
            end_date=date(2026, 8, 31),
            user=self.user,
        )
        self.run = create_run(period=self.period, user=self.user)
        seed_standard_components(self.company)
        structure = SalaryStructure.objects.create(
            company=self.company,
            code='APPR83',
            name='Approval Structure',
        )
        basic = SalaryComponent.objects.get(company=self.company, component_code='BASIC')
        SalaryStructureLine.objects.create(
            structure=structure,
            component=basic,
            calculation_type=CalculationType.PERCENTAGE,
            percent=Decimal('100.0000'),
            display_order=10,
        )
        EmployeeSalaryAssignment.objects.create(
            employee=self.employee,
            salary_structure=structure,
            effective_from=date(2026, 1, 1),
            gross_salary=Decimal('30000.00'),
            ctc=Decimal('360000.00'),
            created_by=self.user,
            updated_by=self.user,
        )
        # Grant workflow permissions to primary test user (Admin-equivalent).
        for codename in ('review_payrollrun', 'approve_payrollrun', 'lock_payrollrun'):
            self.user.user_permissions.add(_perm(codename))
        self.hr_user = get_user_model().objects.create_user(
            username='hrreviewer',
            password='TestPassword123!',
        )
        Group.objects.get(name='HR').user_set.add(self.hr_user)
        self.hr_user.user_permissions.add(_perm('view_payrollrun'))

        self.viewer = get_user_model().objects.create_user(
            username='viewer83',
            password='TestPassword123!',
        )
        Group.objects.get(name='Viewer').user_set.add(self.viewer)
        self.viewer.user_permissions.add(_perm('view_payrollrun'))

        self.superuser = get_user_model().objects.create_superuser(
            username='super83',
            email='super83@example.com',
            password='TestPassword123!',
        )

    def _calculate(self):
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        return self.run

    def test_valid_full_sequence(self):
        self._calculate()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)

        mark_reviewed(self.run, user=self.hr_user, remarks='Looks good')
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.REVIEWED)

        approve_run(self.run, user=self.user, remarks='Approved for lock')
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.APPROVED)

        lock_run(self.run, user=self.user, remarks='Final')
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.LOCKED)
        self.assertTrue(self.run.is_locked)

    def test_invalid_skip_transitions(self):
        self._calculate()
        with self.assertRaises(InvalidTransitionError):
            approve_run(self.run, user=self.user)
        with self.assertRaises(InvalidTransitionError):
            lock_run(self.run, user=self.user)
        # Draft → Reviewed not allowed
        draft = create_run(period=self.period, user=self.user)
        with self.assertRaises(InvalidTransitionError):
            mark_reviewed(draft, user=self.hr_user)

    def test_unauthorized_review_approval_lock(self):
        self._calculate()
        with self.assertRaises(PermissionDenied):
            mark_reviewed(self.run, user=self.viewer)
        mark_reviewed(self.run, user=self.hr_user)
        self.run.refresh_from_db()
        with self.assertRaises(PermissionDenied):
            approve_run(self.run, user=self.hr_user)
        with self.assertRaises(PermissionDenied):
            approve_run(self.run, user=self.viewer)
        approve_run(self.run, user=self.user)
        self.run.refresh_from_db()
        with self.assertRaises(PermissionDenied):
            lock_run(self.run, user=self.hr_user)
        with self.assertRaises(PermissionDenied):
            lock_run(self.run, user=self.viewer)

    def test_approval_blocked_when_errors(self):
        for assignment in self.employee.salary_assignments.all():
            assignment.soft_delete()
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.INCOMPLETE)
        self.assertTrue(self.run.calculation_errors)
        with self.assertRaises(InvalidTransitionError):
            mark_reviewed(self.run, user=self.hr_user)

        # Calculated with injected errors also blocked
        self.run.status = PayrollRunStatus.CALCULATED
        self.run.calculation_errors = [{'employee_code': 'X', 'error': 'boom'}]
        self.run.save(update_fields=['status', 'calculation_errors'])
        with self.assertRaises(RunNotReadyError):
            mark_reviewed(self.run, user=self.hr_user)

    def test_locked_recalc_rejected(self):
        self._calculate()
        mark_reviewed(self.run, user=self.hr_user)
        approve_run(self.run, user=self.user)
        lock_run(self.run, user=self.user)
        with self.assertRaises(LockedRunError):
            calculate_run(self.run, user=self.user)

    def test_locked_result_update_rejected(self):
        self._calculate()
        mark_reviewed(self.run, user=self.hr_user)
        approve_run(self.run, user=self.user)
        lock_run(self.run, user=self.user)
        result = PayrollResult.objects.get(run=self.run)
        result.net_salary = Decimal('1.00')
        with self.assertRaises(ValidationError):
            result.save()
        component = result.components.first()
        component.amount = Decimal('1.00')
        with self.assertRaises(ValidationError):
            component.save()
        with self.assertRaises(ValidationError):
            result.delete()
        with self.assertRaises(ValidationError):
            self.run.delete()

    def test_duplicate_approval_and_lock_rejected(self):
        self._calculate()
        mark_reviewed(self.run, user=self.hr_user)
        approve_run(self.run, user=self.user)
        with self.assertRaises(InvalidTransitionError):
            approve_run(self.run, user=self.user)
        lock_run(self.run, user=self.user)
        with self.assertRaises(InvalidTransitionError):
            lock_run(self.run, user=self.user)
        with self.assertRaises(InvalidTransitionError):
            mark_reviewed(self.run, user=self.hr_user)

    def test_audit_per_transition(self):
        self._calculate()
        mark_reviewed(self.run, user=self.hr_user, remarks='HR ok')
        approve_run(self.run, user=self.user, remarks='Admin ok')
        lock_run(self.run, user=self.user, remarks='Lock it')

        review_log = PayrollAuditLog.objects.filter(run=self.run, action='run_review').first()
        approve_log = PayrollAuditLog.objects.filter(run=self.run, action='run_approve').first()
        lock_log = PayrollAuditLog.objects.filter(run=self.run, action='run_lock').first()
        self.assertIsNotNone(review_log)
        self.assertIsNotNone(approve_log)
        self.assertIsNotNone(lock_log)
        self.assertEqual(review_log.details['previous_status'], PayrollRunStatus.CALCULATED)
        self.assertEqual(review_log.details['new_status'], PayrollRunStatus.REVIEWED)
        self.assertEqual(review_log.details['remarks'], 'HR ok')
        self.assertEqual(review_log.user_id, self.hr_user.pk)
        self.assertEqual(approve_log.details['previous_status'], PayrollRunStatus.REVIEWED)
        self.assertEqual(lock_log.details['new_status'], PayrollRunStatus.LOCKED)

    def test_transaction_rollback_when_audit_fails(self):
        self._calculate()
        with patch(
            'apps.payroll.services.workflow.write_status_transition_audit',
            side_effect=RuntimeError('audit down'),
        ):
            with self.assertRaises(RuntimeError):
                mark_reviewed(self.run, user=self.hr_user)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        self.assertFalse(
            PayrollAuditLog.objects.filter(run=self.run, action='run_review').exists()
        )

    def test_select_for_update_used_on_transition(self):
        self._calculate()
        with patch.object(
            PayrollRun.objects,
            'select_for_update',
            wraps=PayrollRun.objects.select_for_update,
        ) as mocked:
            mark_reviewed(self.run, user=self.hr_user)
            mocked.assert_called()

    def test_reopen_superuser_only_with_reason(self):
        self._calculate()
        mark_reviewed(self.run, user=self.hr_user)
        approve_run(self.run, user=self.user)
        lock_run(self.run, user=self.user)
        with self.assertRaises(PermissionDenied):
            reopen_run(self.run, user=self.user, remarks='need fix')
        with self.assertRaises(InvalidTransitionError):
            reopen_run(self.run, user=self.superuser, remarks='')
        reopen_run(self.run, user=self.superuser, remarks='Correct bank file')
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.CALCULATED)
        self.assertTrue(self.run.is_calculable)
        calculate_run(self.run, user=self.superuser)
        self.run.refresh_from_db()
        self.assertIn(self.run.status, {
            PayrollRunStatus.CALCULATED,
            PayrollRunStatus.INCOMPLETE,
        })
        reopen_log = PayrollAuditLog.objects.filter(run=self.run, action='run_reopen').first()
        self.assertIsNotNone(reopen_log)
        self.assertEqual(reopen_log.details['remarks'], 'Correct bank file')
        self.assertTrue(reopen_log.details.get('reopened_by_superuser'))
        self.assertEqual(reopen_log.details.get('new_status'), PayrollRunStatus.CALCULATED)

    def test_workflow_ui_buttons_and_audit_trail(self):
        self._calculate()
        self.client.force_login(self.user)
        detail = self.client.get(reverse('payroll:run_detail', args=[self.run.pk]))
        self.assertContains(detail, 'Review')
        self.assertContains(detail, 'Audit trail')

        review = self.client.post(
            reverse('payroll:run_review', args=[self.run.pk]),
            {'remarks': 'UI review'},
        )
        self.assertEqual(review.status_code, 302)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.REVIEWED)

        detail = self.client.get(reverse('payroll:run_detail', args=[self.run.pk]))
        self.assertContains(detail, 'Approve')
        self.assertNotContains(detail, '>Recalculate<')

        approve = self.client.post(reverse('payroll:run_approve', args=[self.run.pk]))
        self.assertEqual(approve.status_code, 302)
        lock = self.client.post(
            reverse('payroll:run_lock', args=[self.run.pk]),
            {'remarks': 'UI lock'},
        )
        self.assertEqual(lock.status_code, 302)
        self.run.refresh_from_db()
        self.assertEqual(self.run.status, PayrollRunStatus.LOCKED)

        detail = self.client.get(reverse('payroll:run_detail', args=[self.run.pk]))
        self.assertContains(detail, 'run_review')
        self.assertContains(detail, 'run_lock')
        self.assertContains(detail, 'Calculation disabled')

    def test_unauthorized_ui_review_returns_403(self):
        self._calculate()
        self.client.force_login(self.viewer)
        response = self.client.post(reverse('payroll:run_review', args=[self.run.pk]))
        self.assertEqual(response.status_code, 403)


class PayrollReportsPayslipSprint84Tests(PayrollTestMixin, TestCase):
    """Sprint 8.4 snapshot reporting, export, and payslip coverage."""

    def setUp(self):
        super().setUp()
        self.branch = Branch.objects.create(
            company=self.company, branch_name='HQ', code='HQ',
        )
        self.department = Department.objects.create(
            company=self.company, name='Ops', code='OPS',
        )
        self.employee.branch = self.branch
        self.employee.department = self.department
        self.employee.save(update_fields=['branch', 'department'])
        for codename in ('review_payrollrun', 'approve_payrollrun', 'lock_payrollrun'):
            self.user.user_permissions.add(_perm(codename))

        self.period = create_period(
            company=self.company, month=8, year=2026,
            start_date=date(2026, 8, 1), end_date=date(2026, 8, 31),
            user=self.user,
        )
        self.run = create_run(period=self.period, user=self.user)
        seed_standard_components(self.company)
        self.structure = SalaryStructure.objects.create(
            company=self.company, code='RPT84', name='Report Structure',
        )
        basic = SalaryComponent.objects.get(
            company=self.company, component_code='BASIC',
        )
        SalaryStructureLine.objects.create(
            structure=self.structure, component=basic,
            calculation_type=CalculationType.PERCENTAGE,
            percent=Decimal('100.0000'), display_order=10,
        )
        self.assignment = EmployeeSalaryAssignment.objects.create(
            employee=self.employee, salary_structure=self.structure,
            effective_from=date(2026, 1, 1), gross_salary=Decimal('30000.00'),
            ctc=Decimal('360000.00'), created_by=self.user, updated_by=self.user,
        )

    def _filters(self, visibility='all', **kwargs):
        return ReportFilters(
            company_ids=kwargs.get('company_ids', ()),
            period_ids=kwargs.get('period_ids', ()),
            run_ids=kwargs.get('run_ids', ()),
            branch_ids=kwargs.get('branch_ids', ()),
            department_ids=kwargs.get('department_ids', ()),
            visibility=visibility, user=kwargs.get('user', self.user),
        )

    def _calculate(self):
        calculate_run(self.run, user=self.user)
        self.run.refresh_from_db()
        return self.run

    def _locked_run_with_results(self):
        if self.run.status == PayrollRunStatus.DRAFT:
            self._calculate()
        mark_reviewed(self.run, user=self.user)
        approve_run(self.run, user=self.user)
        lock_run(self.run, user=self.user)
        self.run.refresh_from_db()
        return self.run

    def test_register_totals_match_db(self):
        self._locked_run_with_results()
        totals = aggregate_result_totals(results_queryset(self._filters('final')))
        expected = PayrollResult.objects.filter(run=self.run).aggregate(
            gross=Sum('gross'),
            total_earnings=Sum('total_earnings'),
            total_deductions=Sum('total_deductions'),
            net_salary=Sum('net_salary'),
        )
        self.assertEqual(totals['employee_count'], 1)
        for name, value in expected.items():
            self.assertEqual(totals[name], value)

    def test_earnings_deductions_reconciliation(self):
        self._locked_run_with_results()
        reconciliation = reconcile_earnings_deductions(self._filters('final'))
        self.assertEqual(reconciliation['earnings_difference'], Decimal('0.00'))
        self.assertEqual(reconciliation['deductions_difference'], Decimal('0.00'))

    def test_branch_department_filters(self):
        self._locked_run_with_results()
        branch_rows, _ = branch_summary(
            self._filters('final', branch_ids=(self.branch.pk,))
        )
        department_rows, _ = department_summary(
            self._filters('final', department_ids=(self.department.pk,))
        )
        self.assertEqual([row.name for row in branch_rows], ['HQ'])
        self.assertEqual([row.name for row in department_rows], ['Ops'])
        self.assertEqual(results_queryset(
            self._filters('final', branch_ids=(self.branch.pk,))
        ).count(), 1)

    def test_historical_snapshot_after_master_change(self):
        self._locked_run_with_results()
        result = PayrollResult.objects.get(run=self.run, employee_id=self.employee.pk)
        original_gross = result.gross
        self.assignment.gross_salary = Decimal('90000.00')
        self.assignment.save(update_fields=['gross_salary'])
        dataset = build_payslip_dataset(result)
        self.assertEqual(dataset['gross'], original_gross)
        self.assertEqual(
            aggregate_result_totals(results_queryset(self._filters('final')))['gross'],
            original_gross,
        )

    def test_archived_employee_in_historical_reports(self):
        self._locked_run_with_results()
        result = PayrollResult.objects.get(run=self.run, employee_id=self.employee.pk)
        self.employee.soft_delete()
        self.assertEqual(results_queryset(self._filters('final')).count(), 1)
        self.assertTrue(build_payslip_dataset(result)['employee']['is_archived'])
        self.assertEqual(department_summary(self._filters('final'))[0][0].name, 'Ops')

    def test_draft_vs_locked_visibility(self):
        self._calculate()
        self.assertEqual(results_queryset(self._filters('final')).count(), 0)
        self.assertEqual(results_queryset(self._filters('draft')).count(), 1)
        self._locked_run_with_results()
        self.assertEqual(results_queryset(self._filters('final')).count(), 1)

    def test_permission_company_filtering(self):
        self._locked_run_with_results()
        unprivileged = get_user_model().objects.create_user(
            username='no-report-access', password='TestPassword123!',
        )
        self.assertEqual(companies_visible_to_user(unprivileged), [])
        self.assertEqual(results_queryset(self._filters('all', user=unprivileged)).count(), 0)

        other_company = Company.objects.create(
            client=self.client_record, company_name='Other Payroll Co',
        )
        other_employee = Employee.objects.create(
            company=other_company, employee_code='EMP8402', first_name='Other',
            date_of_joining=date(2026, 1, 1), basic_salary=Decimal('1.00'),
            auto_generate_code=False,
        )
        other_period = create_period(
            company=other_company, month=8, year=2026,
            start_date=date(2026, 8, 1), end_date=date(2026, 8, 31),
            user=self.user,
        )
        other_run = create_run(period=other_period, user=self.user)
        PayrollResult.objects.create(run=other_run, employee=other_employee)
        filtered = results_queryset(
            self._filters('all', company_ids=(self.company.pk,))
        )
        self.assertEqual(filtered.count(), 1)
        self.assertTrue(all(result.run.company_id == self.company.pk for result in filtered))

    def test_excel_export_content_and_totals(self):
        self._locked_run_with_results()
        response = export_payroll_register(self._filters('final'))
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        sheet = load_workbook(BytesIO(response.content), data_only=True).active
        self.assertEqual(sheet.max_column, 14)
        totals = aggregate_result_totals(results_queryset(self._filters('final')))
        self.assertEqual(sheet.cell(sheet.max_row, 1).value, 'Totals')
        self.assertEqual(sheet.cell(sheet.max_row, 11).value, totals['gross'])
        self.assertEqual(sheet.cell(sheet.max_row, 14).value, totals['net_salary'])

    def test_payslip_dataset_ordering_rounding_and_draft_watermark(self):
        self._calculate()
        result = PayrollResult.objects.get(run=self.run, employee_id=self.employee.pk)
        draft = build_payslip_dataset(result)
        self.assertEqual(draft['watermark'], 'DRAFT')
        self.assertEqual(
            [line['code'] for line in draft['earnings']],
            sorted(line['code'] for line in draft['earnings']),
        )
        self.assertTrue(all(
            line['amount'].as_tuple().exponent == -2 for line in draft['earnings']
        ))
        self._locked_run_with_results()
        result.refresh_from_db()
        self.assertIsNone(build_payslip_dataset(result)['watermark'])

    def test_empty_run_report(self):
        filters = self._filters('all', run_ids=(self.run.pk,))
        totals = aggregate_result_totals(results_queryset(filters))
        self.assertEqual(results_queryset(filters).count(), 0)
        self.assertEqual(totals['employee_count'], 0)
        self.assertEqual(totals['net_salary'], Decimal('0.00'))

    def test_engine_report_page_and_export_views(self):
        self._calculate()
        self.client.force_login(self.user)
        page = self.client.get(
            reverse('payroll:engine_report', args=['payroll-register']),
            {'visibility': 'all'},
        )
        export = self.client.get(
            reverse('payroll:engine_report_export', args=['payroll-register']),
            {'visibility': 'all'},
        )
        self.assertEqual(page.status_code, 200)
        self.assertEqual(export.status_code, 200)
        self.assertEqual(
            export['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_payslip_preview_view(self):
        self._calculate()
        result = PayrollResult.objects.get(run=self.run, employee_id=self.employee.pk)
        self.client.force_login(self.user)
        response = self.client.get(reverse(
            'payroll:payslip_preview', args=[self.run.pk, result.pk],
        ))
        self.assertEqual(response.status_code, 200)
