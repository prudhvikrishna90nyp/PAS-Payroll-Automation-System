"""Snapshot-only Excel exports for payroll engine reports."""
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from apps.payroll.models import ComponentType
from .report_queries import (
    aggregate_component_totals, aggregate_result_totals, branch_summary, company_summary,
    component_queryset, department_summary, load_employees_map, results_queryset,
    run_control_queryset, status_label_for_display,
)

REPORT_KEYS = [
    "payroll_register", "earnings_register", "deductions_register", "net_salary_register",
    "employee_payroll_detail", "department_payroll_summary", "branch_payroll_summary",
    "company_payroll_summary", "payroll_run_control",
]
EXPORT_MAP = {key: f"export_{key}" for key in REPORT_KEYS}


def workbook_response(workbook, filename):
    stream = BytesIO()
    workbook.save(stream)
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _workbook(title, headers):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return wb, ws


def _name(instance, attribute):
    return getattr(instance, attribute, "") if instance else ""


def _result_rows(filters):
    results = list(results_queryset(filters))
    employees = load_employees_map({result.employee_id for result in results})
    return results, employees


def _append_total(ws, values):
    ws.append(["Totals", *values])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)


def export_payroll_register(filters):
    headers = ["Company", "Period", "Run#", "Status", "Emp Code", "Emp Name", "Branch", "Department", "Present", "LOP", "Gross", "Earnings", "Deductions", "Net"]
    wb, ws = _workbook("Payroll Register", headers)
    results, employees = _result_rows(filters)
    for result in results:
        employee, run, period = employees.get(result.employee_id), result.run, result.run.period
        ws.append([run.company.company_name, f"{period.month:02d}/{period.year}", run.run_number, status_label_for_display(run.status), getattr(employee, "employee_code", ""), getattr(employee, "full_name", ""), _name(getattr(employee, "branch", None), "branch_name"), _name(getattr(employee, "department", None), "name"), result.present_days, result.lop_days, result.gross, result.total_earnings, result.total_deductions, result.net_salary])
    total = aggregate_result_totals(results_queryset(filters))
    # 14 headers: Totals label + 9 blank non-money columns + 4 monetary totals.
    _append_total(ws, [""] * 9 + [
        total["gross"], total["total_earnings"], total["total_deductions"], total["net_salary"],
    ])
    return workbook_response(wb, "payroll_register.xlsx")


def _component_register(filters, component_type, filename, title):
    wb, ws = _workbook(title, ["Company", "Period", "Run#", "Status", "Emp Code", "Emp Name", "Component Code", "Component Name", "Amount"])
    components = list(component_queryset(filters, component_type))
    employees = load_employees_map({item.result.employee_id for item in components})
    for item in components:
        result, run, employee = item.result, item.result.run, employees.get(item.result.employee_id)
        ws.append([run.company.company_name, f"{run.period.month:02d}/{run.period.year}", run.run_number, status_label_for_display(run.status), getattr(employee, "employee_code", ""), getattr(employee, "full_name", ""), item.component_code, item.component_name, item.amount])
    _append_total(ws, ["", "", "", "", "", "", "", aggregate_component_totals(component_queryset(filters, component_type))["amount"]])
    return workbook_response(wb, filename)


def export_earnings_register(filters):
    return _component_register(filters, ComponentType.EARNING, "earnings_register.xlsx", "Earnings Register")


def export_deductions_register(filters):
    return _component_register(filters, ComponentType.DEDUCTION, "deductions_register.xlsx", "Deductions Register")


def export_net_salary_register(filters):
    wb, ws = _workbook("Net Salary Register", ["Company", "Period", "Run#", "Status", "Emp Code", "Emp Name", "Gross", "Deductions", "Net"])
    results, employees = _result_rows(filters)
    for result in results:
        run, employee = result.run, employees.get(result.employee_id)
        ws.append([run.company.company_name, f"{run.period.month:02d}/{run.period.year}", run.run_number, status_label_for_display(run.status), getattr(employee, "employee_code", ""), getattr(employee, "full_name", ""), result.gross, result.total_deductions, result.net_salary])
    totals = aggregate_result_totals(results_queryset(filters))
    _append_total(ws, ["", "", "", "", "", "", totals["gross"], totals["total_deductions"], totals["net_salary"]])
    return workbook_response(wb, "net_salary_register.xlsx")


def export_employee_payroll_detail(filters):
    wb, ws = _workbook("Employee Payroll Detail", ["Company", "Period", "Run#", "Emp Code", "Emp Name", "Present", "Absent", "LOP", "Overtime", "Gross", "Earnings", "Deductions", "Net", "CTC"])
    results, employees = _result_rows(filters)
    for result in results:
        run, employee = result.run, employees.get(result.employee_id)
        ws.append([run.company.company_name, f"{run.period.month:02d}/{run.period.year}", run.run_number, getattr(employee, "employee_code", ""), getattr(employee, "full_name", ""), result.present_days, result.absent_days, result.lop_days, result.overtime_hours, result.gross, result.total_earnings, result.total_deductions, result.net_salary, result.ctc_snapshot or Decimal("0.00")])
    totals = aggregate_result_totals(results_queryset(filters))
    _append_total(ws, ["", "", "", "", "", "", "", "", "", totals["gross"], totals["total_earnings"], totals["total_deductions"], totals["net_salary"], totals["ctc_snapshot"]])
    return workbook_response(wb, "employee_payroll_detail.xlsx")


def _summary_export(rows, title, filename, label):
    wb, ws = _workbook(title, [label, "Employees", "Gross", "Earnings", "Deductions", "Net"])
    for row in rows:
        ws.append([row.name, row.employee_count, row.gross, row.total_earnings, row.total_deductions, row.net_salary])
    totals = {"employee_count": sum(row.employee_count for row in rows), "gross": sum((row.gross for row in rows), Decimal("0.00")), "total_earnings": sum((row.total_earnings for row in rows), Decimal("0.00")), "total_deductions": sum((row.total_deductions for row in rows), Decimal("0.00")), "net_salary": sum((row.net_salary for row in rows), Decimal("0.00"))}
    _append_total(ws, [totals["employee_count"], totals["gross"], totals["total_earnings"], totals["total_deductions"], totals["net_salary"]])
    return workbook_response(wb, filename)


def export_department_summary(filters):
    rows, _ = department_summary(filters)
    return _summary_export(rows, "Department Summary", "department_payroll_summary.xlsx", "Department")

def export_branch_summary(filters):
    rows, _ = branch_summary(filters)
    return _summary_export(rows, "Branch Summary", "branch_payroll_summary.xlsx", "Branch")

def export_company_summary(filters):
    rows, _ = company_summary(filters)
    return _summary_export(rows, "Company Summary", "company_payroll_summary.xlsx", "Company")


def export_run_control(filters):
    wb, ws = _workbook("Payroll Run Control", ["Company", "Period", "Run#", "Status", "Results", "Gross", "Earnings", "Deductions", "Net"])
    for run in run_control_queryset(filters):
        totals = aggregate_result_totals(run.results.all())
        ws.append([run.company.company_name, f"{run.period.month:02d}/{run.period.year}", run.run_number, status_label_for_display(run.status), run.results.count(), totals["gross"], totals["total_earnings"], totals["total_deductions"], totals["net_salary"]])
    totals = aggregate_result_totals(results_queryset(filters))
    _append_total(ws, ["", "", "", results_queryset(filters).count(), totals["gross"], totals["total_earnings"], totals["total_deductions"], totals["net_salary"]])
    return workbook_response(wb, "payroll_run_control.xlsx")


export_department_payroll_summary = export_department_summary
export_branch_payroll_summary = export_branch_summary
export_company_payroll_summary = export_company_summary
export_payroll_run_control = export_run_control


ENGINE_EXPORT_BUILDERS = {
    "payroll_register": export_payroll_register, "earnings_register": export_earnings_register,
    "deductions_register": export_deductions_register, "net_salary_register": export_net_salary_register,
    "employee_payroll_detail": export_employee_payroll_detail, "department_payroll_summary": export_department_payroll_summary,
    "branch_payroll_summary": export_branch_payroll_summary, "company_payroll_summary": export_company_payroll_summary,
    "payroll_run_control": export_payroll_run_control,
}
